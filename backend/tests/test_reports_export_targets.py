from io import BytesIO

from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlmodel import Session, create_engine

from app import database as db_module
from app import main as app_main
from app import models
from app.main import ExcelExportPayload, ReportClass, ReportStudent, _resolve_export_targets


def test_resolve_export_targets_falls_back_to_turma_and_horario_when_professor_differs():
    reports = [
        ReportClass(
            turma="Terça e Quinta",
            turmaCodigo="INF-A",
            horario="18:30",
            professor="Professor João",
            nivel="Iniciante",
            hasLog=True,
            alunos=[
                ReportStudent(
                    id="1",
                    student_uid=None,
                    nome="Ana Silva",
                    presencas=1,
                    faltas=0,
                    justificativas=0,
                    frequencia=100.0,
                    historico={},
                )
            ],
        )
    ]

    payload = ExcelExportPayload(
        month="2026-03",
        classes=[
            {
                "turma": "Terça e Quinta",
                "horario": "18:30",
                "professor": "Professor A",
            }
        ],
    )

    selected = _resolve_export_targets(payload, reports)

    assert len(selected) == 1
    assert selected[0].professor == "Professor João"
    assert selected[0].turmaCodigo == "INF-A"


def test_excel_export_creates_separate_sheets_for_each_class(tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    test_db_path = tmp_path / "reports.db"
    test_engine = create_engine(
        f"sqlite:///{test_db_path}",
        connect_args={"check_same_thread": False},
    )

    app_main.DATA_DIR = str(data_dir)
    db_module.engine = test_engine
    db_module.create_db_and_tables()

    with Session(test_engine) as session:
        unit = models.ImportUnit(name="Unidade Teste")
        session.add(unit)
        session.commit()
        session.refresh(unit)

        class_a = models.ImportClass(
            unit_id=unit.id,
            codigo="INF-A",
            turma_label="Terça e Quinta",
            horario="18:30",
            professor="Professor A",
            nivel="Iniciante",
            capacidade=20,
        )
        class_b = models.ImportClass(
            unit_id=unit.id,
            codigo="INF-B",
            turma_label="Quarta e Sexta",
            horario="19:15",
            professor="Professor B",
            nivel="Intermediário",
            capacidade=20,
        )
        session.add(class_a)
        session.add(class_b)
        session.commit()
        session.refresh(class_a)
        session.refresh(class_b)

        session.add(models.ImportStudent(class_id=class_a.id, nome="Ana Silva"))
        session.add(models.ImportStudent(class_id=class_b.id, nome="Bruno Lima"))
        session.commit()

    with TestClient(app_main.app) as client:
        response = client.post(
            "/reports/excel-file",
            json={
                "month": "2026-03",
                "classes": [
                    {
                        "turmaCodigo": "INF-A",
                        "turma": "Terça e Quinta",
                        "horario": "18:30",
                        "professor": "Professor A",
                    },
                    {
                        "turmaCodigo": "INF-B",
                        "turma": "Quarta e Sexta",
                        "horario": "19:15",
                        "professor": "Professor B",
                    },
                ],
            },
        )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert sorted(workbook.sheetnames) == ["18-30|19-15", "19-15|20-00"]

    first_sheet = workbook["18-30|19-15"]
    second_sheet = workbook["19-15|20-00"]

    assert first_sheet["A7"].value == "Ana Silva"
    assert second_sheet["A7"].value == "Bruno Lima"


def test_get_reports_ignores_students_outside_class_roster(tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    test_db_path = tmp_path / "reports.db"
    test_engine = create_engine(
        f"sqlite:///{test_db_path}",
        connect_args={"check_same_thread": False},
    )

    app_main.DATA_DIR = str(data_dir)
    db_module.engine = test_engine
    db_module.create_db_and_tables()

    with Session(test_engine) as session:
        unit = models.ImportUnit(name="Unidade Teste")
        session.add(unit)
        session.commit()
        session.refresh(unit)

        class_a = models.ImportClass(
            unit_id=unit.id,
            codigo="jtg01",
            turma_label="Terça e Quinta",
            horario="08:00",
            professor="Jefferson",
            nivel="Nível 4 B",
            capacidade=20,
        )
        session.add(class_a)
        session.commit()
        session.refresh(class_a)

        session.add(models.ImportStudent(class_id=class_a.id, nome="Aluno Da Turma"))
        session.commit()

        session.add(
            models.AttendanceLog(
                turma_codigo="jtg01",
                turma_label="Terça e Quinta",
                horario="08:00",
                professor="Jefferson",
                mes="2026-04",
                saved_at="2026-04-01T00:00:00Z",
                registros_json=(
                    "["
                    '{"aluno_nome":"Aluno Da Turma","attendance":{"2026-04-01":"Presente"}},'
                    '{"aluno_nome":"Aluno De Outra Turma","attendance":{"2026-04-01":"Presente"}}'
                    "]"
                ),
            )
        )
        session.commit()

        reports = app_main.get_reports(month="2026-04", session=session)

    assert len(reports) == 1
    assert [aluno.nome for aluno in reports[0].alunos] == ["Aluno da Turma"]