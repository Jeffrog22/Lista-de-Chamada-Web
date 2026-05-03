from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Form, Response
from sqlmodel import Session, select, func
from app.database import create_db_and_tables, migrate_db, get_session, engine
from app import crud, models
from app.models import AttendanceLog, AcademicCalendarState, PoolLog, ExclusionRecord
from typing import List, Optional, Dict, Any, Tuple
from contextlib import asynccontextmanager
import os
import json
import re
import uuid
import math
from datetime import datetime, timedelta, date
from threading import RLock
import unicodedata
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from pydantic import BaseModel, Field, ConfigDict
import csv
from io import StringIO, BytesIO
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from app.etl.import_excel import import_from_excel
from app.auth import get_password_hash, create_access_token, authenticate_user, get_current_user
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

@asynccontextmanager
async def lifespan(_: FastAPI):
    create_db_and_tables()
    migrate_db()
    os.makedirs(DATA_DIR, exist_ok=True)
    yield


app = FastAPI(title="Lista-de-Chamada - API", lifespan=lifespan)

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
load_dotenv(os.path.join(BASE_DIR, ".env"))
load_dotenv(os.path.join(BASE_DIR, "backend", ".env"))

DATA_DIR = (
    os.getenv("APP_DATA_DIR", "").strip()
    or os.getenv("DATA_DIR", "").strip()
    or os.path.join(BASE_DIR, "data")
)
EXCLUSIONS_FILE_LOCK = RLock()
ACADEMIC_CALENDAR_FILE_LOCK = RLock()

ENV_NAME = os.getenv("ENV_NAME", "").strip()
UNIT_NAME = os.getenv("UNIT_NAME", "").strip()
ACCESS_MODE = os.getenv("ACCESS_MODE", "unit").strip().lower()

cors_origins_raw = os.getenv(
    "CORS_ORIGINS",
    "http://localhost:3000,http://localhost:5173,https://chamadabelavista.pages.dev",
)


def _normalize_origin(origin: str) -> str:
    return origin.strip().strip("\"'").rstrip("/")


origins = [
    _normalize_origin(origin)
    for origin in cors_origins_raw.split(",")
    if _normalize_origin(origin)
]
for mandatory_origin in [
    "https://chamadabelavista.pages.dev",
    "https://chamadasaomatheus.pages.dev",
    "https://chamadavila.pages.dev",
]:
    normalized_mandatory = _normalize_origin(mandatory_origin)
    if normalized_mandatory and normalized_mandatory not in origins:
        origins.append(normalized_mandatory)

cors_origin_regex = os.getenv(
    "CORS_ORIGIN_REGEX",
    r"^https://.*\.(vercel\.app|netlify\.app|pages\.dev)$",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_origin_regex=cors_origin_regex if cors_origin_regex else None,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
def health():
    return {"status": "ok"}


def _normalize_unit_name(value: str) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFD", raw)
    folded = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    if folded in {"bela vista", "piscina bela vista"}:
        return "bela vista"
    return folded


def _validate_unit_for_environment(typed_unit: Optional[str]) -> None:
    # UNIT_NAME vazio mantém comportamento legado (sem bloqueio por ambiente).
    if not UNIT_NAME:
        return

    if not typed_unit:
        raise HTTPException(
            status_code=400,
            detail=f"Unidade obrigatória para este ambiente. Unidade oficial: {UNIT_NAME}",
        )

    if _normalize_unit_name(typed_unit) != _normalize_unit_name(UNIT_NAME):
        raise HTTPException(
            status_code=400,
            detail=f"Login bloqueado: este ambiente aceita apenas a unidade '{UNIT_NAME}'.",
        )


@app.get("/environment")
def environment_info():
    return {
        "env_name": ENV_NAME,
        "unit_name": UNIT_NAME,
        "access_mode": ACCESS_MODE or "unit",
    }

class PoolLogEntryModel(BaseModel):
    data: str
    turmaCodigo: str = ""
    turmaLabel: str = ""
    horario: str = ""
    professor: str = ""
    clima1: str
    clima2: str
    statusAula: str
    nota: str
    tipoOcorrencia: str
    tempExterna: Optional[str] = ""
    tempPiscina: Optional[str] = ""
    cloroPpm: Optional[float] = None

class AttendanceLogItem(BaseModel):
    aluno_nome: str
    attendance: Dict[str, str]
    justifications: Optional[Dict[str, str]] = None
    notes: Optional[List[str]] = None

class AttendanceLogPayload(BaseModel):
    turmaCodigo: str = ""
    turmaLabel: str = ""
    horario: str = ""
    professor: str = ""
    mes: str = ""
    clientSavedAt: Optional[str] = ""
    clientMutationId: Optional[int] = None
    registros: List[AttendanceLogItem]

class AttendanceSyncProbePayload(BaseModel):
    turmaCodigo: str = ""
    turmaLabel: str = ""
    horario: str = ""
    professor: str = ""
    mes: str = ""

class JustificationLogEntry(BaseModel):
    aluno_nome: str
    data: str
    motivo: str
    turmaCodigo: str = ""
    turmaLabel: str = ""
    horario: str = ""
    professor: str = ""

class ExclusionEntry(BaseModel):
    id: Optional[str] = None
    student_uid: Optional[str] = None
    nome: Optional[str] = None
    turma: Optional[str] = None
    turmaCodigo: Optional[str] = None
    horario: Optional[str] = None
    professor: Optional[str] = None
    dataExclusao: Optional[str] = None
    motivo_exclusao: Optional[str] = None

    model_config = ConfigDict(extra="allow")


class ExclusionsBulkPayload(BaseModel):
    items: List[ExclusionEntry] = Field(default_factory=list)
    replace: bool = False


class ExclusionsRecoverPayload(BaseModel):
    backup_file: Optional[str] = None
    force: bool = False
    merge: bool = True
    expected_min_items: int = 1

class ReportStudent(BaseModel):
    id: str
    student_uid: Optional[str] = None
    nome: str
    presencas: int
    faltas: int
    justificativas: int
    frequencia: float
    historico: Dict[str, str]
    justifications: Dict[str, str] = Field(default_factory=dict)
    notes: List[str] = Field(default_factory=list)
    anotacoes: Optional[str] = None

class ReportClass(BaseModel):
    turma: str
    turmaCodigo: str = ""
    horario: str
    professor: str
    nivel: str
    hasLog: bool = False
    alunos: List[ReportStudent]

class ReportsFilterOut(BaseModel):
    turmas: List[str]
    horarios: List[str]
    professores: List[str]
    meses: List[str]
    anos: List[str]


# --- Statistics models (per-student retention & per-level permanence) ---
class LevelHistoryOut(BaseModel):
    nivel: str
    firstDate: Optional[str]
    lastDate: Optional[str]
    days: int
    presencas: int
    faltas: int
    justificativas: int
    frequencia: float

class StudentStatisticsOut(BaseModel):
    id: Optional[str] = None
    nome: str
    firstPresence: Optional[str] = None
    lastPresence: Optional[str] = None
    exclusionDate: Optional[str] = None
    retentionDays: int = 0
    currentNivel: Optional[str] = None
    levels: List[LevelHistoryOut] = []

class ExportClassSelection(BaseModel):
    turmaCodigo: Optional[str] = None
    turma: str
    horario: str
    professor: str

class ExcelExportPayload(BaseModel):
    month: Optional[str] = None
    turma: Optional[str] = None
    horario: Optional[str] = None
    professor: Optional[str] = None
    classes: List[ExportClassSelection] = Field(default_factory=list)


class VacancyExportDetail(BaseModel):
    nivel: str
    lotacao: int
    capacidade: int
    professor: str


class VacancyExportBlock(BaseModel):
    groupKey: str
    periodoLabel: str
    horario: str
    lotacaoHorario: int
    capacidadeHorario: int
    vagasDisponiveis: int
    excesso: int
    rows: List[VacancyExportDetail] = Field(default_factory=list)


class VacancyExportSummary(BaseModel):
    totalCapacidade: int = 0
    totalLotacao: int = 0
    totalVagas: int = 0
    totalExcesso: int = 0


class VacancyExportPayload(BaseModel):
    generatedAt: Optional[str] = None
    summary: VacancyExportSummary
    blocks: List[VacancyExportBlock] = Field(default_factory=list)

class AcademicCalendarSettingsPayload(BaseModel):
    schoolYear: int
    inicioAulas: str
    feriasInvernoInicio: str
    feriasInvernoFim: str
    terminoAulas: str

class AcademicCalendarEventPayload(BaseModel):
    date: str
    type: str
    allDay: Optional[bool] = False
    startTime: Optional[str] = ""
    endTime: Optional[str] = ""
    description: Optional[str] = ""
    teacher: Optional[str] = ""


class PlanningBlockModel(BaseModel):
    id: str
    type: str
    key: str
    label: str
    text: str
    month: Optional[str] = None
    week: Optional[int] = None
    startDay: Optional[int] = None
    endDay: Optional[int] = None


class PlanningFileModel(BaseModel):
    id: str
    sourceName: str
    target: str
    year: int
    blocks: List[PlanningBlockModel] = Field(default_factory=list)
    createdAt: str

def _append_json_list(file_path: str, items: List[Dict[str, Any]]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    payload: List[Dict[str, Any]] = []
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            if not isinstance(payload, list):
                raise RuntimeError(f"Arquivo JSON inválido para append: {file_path} (esperado array)")
        except Exception:
            # Fail-open strategy for runtime persistence: corrupted/partial JSON
            # must not break attendance writes in production.
            try:
                archive_dir = os.path.join(DATA_DIR, "archive")
                os.makedirs(archive_dir, exist_ok=True)
                ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                backup_path = os.path.join(
                    archive_dir,
                    f"{os.path.splitext(os.path.basename(file_path))[0]}_corrupted_{ts}.json",
                )
                if os.path.exists(file_path):
                    with open(file_path, "r", encoding="utf-8") as src, open(
                        backup_path, "w", encoding="utf-8"
                    ) as dst:
                        dst.write(src.read())
            except Exception:
                pass
            payload = []
    payload.extend(items)
    _backup_runtime_json(file_path)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

def _load_json_list(file_path: str) -> List[Dict[str, Any]]:
    if not os.path.exists(file_path):
        return []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return payload if isinstance(payload, list) else []
    except Exception:
        return []

def _backup_runtime_json(file_path: str) -> None:
    try:
        if not os.path.exists(file_path):
            return

        with open(file_path, "r", encoding="utf-8") as f:
            raw = f.read()

        if not raw.strip():
            return

        archive_dir = os.path.join(DATA_DIR, "archive")
        os.makedirs(archive_dir, exist_ok=True)

        base_name = os.path.splitext(os.path.basename(file_path))[0]
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(archive_dir, f"{base_name}_backup_{ts}.json")

        with open(backup_path, "w", encoding="utf-8") as f:
            f.write(raw)
    except Exception:
        # best-effort only
        return

def _save_json_list(file_path: str, items: List[Dict[str, Any]]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    _backup_runtime_json(file_path)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def _weather_snapshots_file() -> str:
    return os.path.join(DATA_DIR, "weatherSnapshots.json")


def _load_weather_snapshots() -> Dict[str, Dict[str, Any]]:
    file_path = _weather_snapshots_file()
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if not isinstance(payload, dict):
            return {}
        return payload
    except Exception:
        return {}


def _save_weather_snapshots(payload: Dict[str, Dict[str, Any]]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    file_path = _weather_snapshots_file()
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _build_weather_snapshot(date_key: str, temp: str, condition: str, condition_code: str) -> Dict[str, Any]:
    return {
        "date": date_key,
        "temp": str(temp or "").strip(),
        "condition": str(condition or "").strip(),
        "conditionCode": str(condition_code or "").strip().lower(),
        "saved_at": datetime.utcnow().isoformat(),
    }


def _compute_weather_temp(minima_txt: str, maxima_txt: str, fallback: str = "26") -> str:
    temp = fallback
    try:
        if minima_txt and maxima_txt:
            temp = str(round((float(minima_txt) + float(maxima_txt)) / 2))
        elif maxima_txt:
            temp = str(round(float(maxima_txt)))
        elif minima_txt:
            temp = str(round(float(minima_txt)))
    except Exception:
        temp = fallback
    return temp

def _academic_calendar_file() -> str:
    return os.path.join(DATA_DIR, "academicCalendar.json")


def _empty_academic_calendar_state() -> Dict[str, Any]:
    return {"settings": None, "events": [], "bankHours": []}


def _normalize_academic_calendar_state(payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return _empty_academic_calendar_state()
    settings = payload.get("settings") if isinstance(payload.get("settings"), dict) else None
    events = payload.get("events") if isinstance(payload.get("events"), list) else []
    bank_hours = payload.get("bankHours") if isinstance(payload.get("bankHours"), list) else []
    return {"settings": settings, "events": events, "bankHours": bank_hours}


def _load_academic_calendar_state_from_db() -> Optional[Dict[str, Any]]:
    try:
        from app.database import engine as _db_engine

        with Session(_db_engine) as db:
            row = db.get(AcademicCalendarState, 1)
            if not row or not str(row.state_json or "").strip():
                return None
            return _normalize_academic_calendar_state(json.loads(row.state_json))
    except Exception:
        return None


def _save_academic_calendar_state_to_db(state: Dict[str, Any]) -> None:
    try:
        from app.database import engine as _db_engine

        with Session(_db_engine) as db:
            row = db.get(AcademicCalendarState, 1)
            payload = json.dumps(_normalize_academic_calendar_state(state), ensure_ascii=False)
            if row:
                row.state_json = payload
                row.updated_at = datetime.utcnow().isoformat()
                db.add(row)
            else:
                db.add(
                    AcademicCalendarState(
                        id=1,
                        state_json=payload,
                        updated_at=datetime.utcnow().isoformat(),
                    )
                )
            db.commit()
    except Exception:
        pass

def _load_academic_calendar_state() -> Dict[str, Any]:
    db_state = _load_academic_calendar_state_from_db()
    if db_state:
        return db_state

    file_path = _academic_calendar_file()
    with ACADEMIC_CALENDAR_FILE_LOCK:
        if not os.path.exists(file_path):
            return _empty_academic_calendar_state()
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            state = _normalize_academic_calendar_state(payload)
            _save_academic_calendar_state_to_db(state)
            return state
        except Exception:
            return _empty_academic_calendar_state()

def _save_academic_calendar_state(state: Dict[str, Any]) -> None:
    normalized_state = _normalize_academic_calendar_state(state)
    _save_academic_calendar_state_to_db(normalized_state)

    os.makedirs(DATA_DIR, exist_ok=True)
    file_path = _academic_calendar_file()
    with ACADEMIC_CALENDAR_FILE_LOCK:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(normalized_state, f, ensure_ascii=False, indent=2)


@app.get("/planning-files")
def get_planning_files() -> List[Dict[str, Any]]:
    return _load_planning_files()


@app.post("/planning-files")
def save_planning_file(file_data: PlanningFileModel) -> Dict[str, Any]:
    files = _load_planning_files()
    files = [entry for entry in files if entry.get("id") != file_data.id]
    entry = file_data.dict()
    files.insert(0, entry)
    _save_planning_files(files)
    return entry


@app.delete("/planning-files/{file_id}")
def delete_planning_file(file_id: str) -> Dict[str, bool]:
    files = _load_planning_files()
    filtered = [entry for entry in files if entry.get("id") != file_id]
    _save_planning_files(filtered)
    return {"ok": True}


def _planning_files_path() -> str:
    return os.path.join(DATA_DIR, "planningFiles.json")


def _load_planning_files() -> List[Dict[str, Any]]:
    return _load_json_list(_planning_files_path())


def _save_planning_files(items: List[Dict[str, Any]]) -> None:
    _save_json_list(_planning_files_path(), items)

def _month_bounds(month: str) -> tuple[str, str]:
    year, month_num = month.split("-")
    start = datetime(int(year), int(month_num), 1)
    if int(month_num) == 12:
        next_month = datetime(int(year) + 1, 1, 1)
    else:
        next_month = datetime(int(year), int(month_num) + 1, 1)
    end = next_month - timedelta(days=1)
    return (start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))

def _hours_from_interval(start_time: str, end_time: str) -> float:
    try:
        start = datetime.strptime(start_time, "%H:%M")
        end = datetime.strptime(end_time, "%H:%M")
        minutes = (end - start).total_seconds() / 60
        if minutes <= 0:
            return 0.0
        return round(minutes / 60, 2)
    except Exception:
        return 0.0

POOL_LOG_COLUMNS = [
    "Data",
    "TurmaCodigo",
    "TurmaLabel",
    "Horario",
    "Professor",
    "Clima 1",
    "Clima 2",
    "Status_aula",
    "Nota",
    "Tipo_ocorrencia",
    "Temp. (C)",
    "Piscina (C)",
    "Cloro (ppm)",
]

def _format_horario(value: str) -> str:
    raw = str(value or "").strip()
    if raw == "":
        return ""
    if ":" in raw:
        parts = raw.split(":")
        if len(parts) >= 2:
            hh = parts[0].zfill(2)[:2]
            mm = parts[1].zfill(2)[:2]
            return f"{hh}:{mm}"
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 3:
        digits = "0" + digits
    if len(digits) >= 4:
        return f"{digits[:2]}:{digits[2:4]}"
    return raw

def _horario_to_minutes(value: str) -> Optional[int]:
    normalized = _format_horario(value)
    if not normalized or ":" not in normalized:
        return None
    try:
        hh, mm = normalized.split(":", 1)
        hour = int(hh)
        minute = int(mm)
    except Exception:
        return None
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None
    return hour * 60 + minute

def _horario_bucket(value: str) -> str:
    minutes = _horario_to_minutes(value)
    if minutes is None:
        return ""
    if minutes < 12 * 60:
        return "manha"
    if minutes < 18 * 60:
        return "tarde"
    return "noite"

def _horario_matches(requested: str, stored: str) -> bool:
    req = _format_horario(requested)
    saved = _format_horario(stored)
    if not req:
        return True
    if not saved:
        return False
    if req == saved:
        return True
    # Compatibilidade retroativa: alguns logs antigos usam horário âncora
    # por período (ex.: 13:00) e não o horário exato da turma (ex.: 16:00).
    req_bucket = _horario_bucket(req)
    saved_bucket = _horario_bucket(saved)
    return bool(req_bucket and saved_bucket and req_bucket == saved_bucket)

def _format_whatsapp(value: Any) -> str:
    raw = str(value or "").strip()
    if raw == "":
        return ""
    if any(ch.isalpha() for ch in raw):
        return ""

    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return ""


def _coerce_numeric_value(value: Optional[str]) -> str | float:
    raw = str(value or "").replace(",", ".").strip()
    if raw == "":
        return ""
    try:
        num = float(raw)
        if math.isfinite(num):
            return num
    except ValueError:
        pass
    return raw


def _format_temperature_output(value: Any, fallback: str = "28") -> str:
    if value is None:
        return fallback
    raw = str(value).strip().replace(",", ".")
    if not raw:
        return fallback
    try:
        num = float(raw)
        if math.isfinite(num):
            return str(num)
    except Exception:
        pass
    return fallback


def _normalize_excel_string(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _normalize_date_key(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        return value.date().isoformat()

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    raw = str(value).strip()
    if not raw:
        return ""

    raw_no_time = raw.replace("T", " ").split(" ")[0]
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw_no_time):
        return raw_no_time

    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", raw_no_time):
        dd, mm, yyyy = raw_no_time.split("/")
        return f"{yyyy}-{mm}-{dd}"

    parsed = pd.to_datetime(raw, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        return parsed.date().isoformat()

    return raw_no_time

def _load_pool_log(file_path: str) -> pd.DataFrame:
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
    else:
        df = pd.DataFrame(columns=POOL_LOG_COLUMNS)

    for col in POOL_LOG_COLUMNS:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype("object")
    return df


def _pool_log_row_from_entry(entry: PoolLogEntryModel) -> Dict[str, Any]:
    cloro_value = "-" if entry.nota in {"feriado", "ponte-feriado", "reuniao"} else ("-" if entry.cloroPpm is None else entry.cloroPpm)
    temp_externa_value = _coerce_numeric_value(entry.tempExterna)
    temp_piscina_value = _coerce_numeric_value(entry.tempPiscina)
    return {
        "Data": _normalize_date_key(entry.data),
        "TurmaCodigo": entry.turmaCodigo or "",
        "TurmaLabel": entry.turmaLabel or "",
        "Horario": _format_horario(entry.horario),
        "Professor": entry.professor or "",
        "Clima 1": entry.clima1,
        "Clima 2": entry.clima2,
        "Status_aula": entry.statusAula,
        "Nota": entry.nota,
        "Tipo_ocorrencia": entry.tipoOcorrencia,
        "Temp. (C)": temp_externa_value,
        "Piscina (C)": temp_piscina_value,
        "Cloro (ppm)": cloro_value,
    }


def _load_pool_log_dataframe() -> pd.DataFrame:
    try:
        from app.database import engine as _db_engine
        from sqlmodel import Session as _DBSession

        with _DBSession(_db_engine) as _db:
            rows = _db.exec(select(PoolLog)).all()
            if rows:
                payload = []
                for row in rows:
                    payload.append({
                        "Data": _normalize_date_key(row.data),
                        "TurmaCodigo": row.turma_codigo or "",
                        "TurmaLabel": row.turma_label or "",
                        "Horario": _format_horario(row.horario),
                        "Professor": row.professor or "",
                        "Clima 1": row.clima1 or "",
                        "Clima 2": row.clima2 or "",
                        "Status_aula": row.status_aula or "",
                        "Nota": row.nota or "",
                        "Tipo_ocorrencia": row.tipo_ocorrencia or "",
                        "Temp. (C)": row.temp_externa or "",
                        "Piscina (C)": row.temp_piscina or "",
                        "Cloro (ppm)": row.cloro_ppm,
                        "saved_at": row.saved_at or "",
                    })
                df = pd.DataFrame(payload)
                for col in [*POOL_LOG_COLUMNS, "saved_at"]:
                    if col not in df.columns:
                        df[col] = ""
                    df[col] = df[col].astype("object")
                return df
    except Exception:
        pass

    file_path = os.path.join(DATA_DIR, "logPiscina.xlsx")
    return _load_pool_log(file_path)

def _pool_log_mask(df: pd.DataFrame, entry: PoolLogEntryModel) -> pd.Series:
    def _norm(value: Any) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        return str(value).strip()

    def _norm_key(value: str) -> str:
        return _normalize_text_fold(_norm(value))

    data_val = _normalize_date_key(entry.data)
    turma_codigo = _norm_key(entry.turmaCodigo)
    turma_label = _norm_key(entry.turmaLabel)
    horario = _format_horario(entry.horario)
    professor = _norm_key(entry.professor)

    mask = df["Data"].apply(_normalize_date_key) == data_val
    if turma_codigo:
        turma_codigo_col = df["TurmaCodigo"].astype(str).map(_norm_key)
        if turma_label:
            # Compatibilidade retroativa: quando o código vier vazio em registros antigos,
            # permite casar pelo label mantendo os demais critérios.
            mask = mask & ((turma_codigo_col == turma_codigo) | (turma_codigo_col == ""))
        else:
            mask = mask & (turma_codigo_col == turma_codigo)
    if turma_label:
        turma_label_col = df["TurmaLabel"].astype(str).map(_norm_key)
        if turma_codigo:
            # Compatibilidade retroativa: quando o label vier vazio em registros antigos,
            # permite casar pelo código mantendo os demais critérios.
            mask = mask & ((turma_label_col == turma_label) | (turma_label_col == ""))
        else:
            mask = mask & (turma_label_col == turma_label)
    if horario:
        mask = mask & df["Horario"].astype(str).apply(lambda value: _horario_matches(horario, value))
    if professor:
        professor_col = df["Professor"].astype(str).map(_norm_key)
        # Compatibilidade retroativa: professor vazio no histórico deve casar
        # com o professor informado na seleção atual.
        mask = mask & ((professor_col == professor) | (professor_col == ""))
    return mask

def _pool_log_row_as_response(row: Dict[str, Any]) -> Dict[str, Any]:
    raw_cloro = row.get("Cloro (ppm)", None)
    cloro_value = None
    if raw_cloro is not None:
        raw_cloro_str = str(raw_cloro).strip().replace(",", ".")
        if raw_cloro_str and raw_cloro_str != "-":
            try:
                parsed_cloro = float(raw_cloro_str)
                if math.isfinite(parsed_cloro):
                    cloro_value = parsed_cloro
            except Exception:
                pass

    return {
        "data": _normalize_date_key(row.get("Data", "")),
        "turmaCodigo": _normalize_excel_string(row.get("TurmaCodigo", "")),
        "turmaLabel": _normalize_excel_string(row.get("TurmaLabel", "")),
        "horario": _normalize_excel_string(row.get("Horario", "")),
        "professor": _normalize_excel_string(row.get("Professor", "")),
        "clima1": _normalize_excel_string(row.get("Clima 1", "")),
        "clima2": _normalize_excel_string(row.get("Clima 2", "")),
        "statusAula": _normalize_excel_string(row.get("Status_aula", "")),
        "nota": _normalize_excel_string(row.get("Nota", "")),
        "tipoOcorrencia": _normalize_excel_string(row.get("Tipo_ocorrencia", "")),
        "tempExterna": _format_temperature_output(row.get("Temp. (C)", None), ""),
        "tempPiscina": _format_temperature_output(row.get("Piscina (C)", None), "28"),
        "cloroPpm": cloro_value,
    }

def _pool_log_meaningful_signature(row: Dict[str, Any]) -> tuple:
    return (
        _normalize_excel_string(row.get("Clima 1", "")),
        _normalize_excel_string(row.get("Clima 2", "")),
        _normalize_excel_string(row.get("Status_aula", "")),
        _normalize_excel_string(row.get("Nota", "")),
        _normalize_excel_string(row.get("Tipo_ocorrencia", "")),
        _format_temperature_output(row.get("Temp. (C)", None), ""),
        _format_temperature_output(row.get("Piscina (C)", None), "28"),
        "" if row.get("Cloro (ppm)", None) is None else _normalize_excel_string(row.get("Cloro (ppm)", None)),
    )

def _select_latest_pool_log_from_rows(rows: pd.DataFrame, requested_horario: str) -> Optional[Dict[str, Any]]:
    if rows.empty:
        return None

    ordered_rows = rows.copy()
    ordered_rows["__horario_minutes"] = ordered_rows["Horario"].astype(str).map(
        lambda value: _horario_to_minutes(_normalize_excel_string(value))
    )
    ordered_rows["__horario_minutes"] = ordered_rows["__horario_minutes"].apply(
        lambda value: value if value is not None else 10**9
    )
    if "saved_at" in ordered_rows.columns:
        ordered_rows["__saved_at_key"] = ordered_rows["saved_at"].astype(str).map(_saved_at_sort_key)
    else:
        ordered_rows["__saved_at_key"] = -1
    ordered_rows = ordered_rows.sort_values(["__horario_minutes", "__saved_at_key"], kind="mergesort")
    ordered_rows = ordered_rows.drop(columns=["__horario_minutes", "__saved_at_key"], errors="ignore")

    requested_minutes = _horario_to_minutes(requested_horario)
    selected: Optional[Dict[str, Any]] = None
    selected_minutes = -1
    selected_index = -1

    for row_index, row in ordered_rows.iterrows():
        row_minutes = _horario_to_minutes(_normalize_excel_string(row.get("Horario", "")))
        if requested_minutes is not None and row_minutes is not None and row_minutes > requested_minutes:
            continue
        current_minutes = row_minutes if row_minutes is not None else -1
        if current_minutes > selected_minutes or (current_minutes == selected_minutes and int(row_index) >= selected_index):
            selected = row.to_dict()
            selected_minutes = current_minutes
            selected_index = int(row_index)

    if selected is not None:
        return selected

    return ordered_rows.iloc[0].to_dict()


def _select_latest_pool_log_for_day(
    df: pd.DataFrame,
    date_value: str,
    requested_horario: str,
    requested_professor: str = "",
    requested_turma_codigo: str = "",
    requested_turma_label: str = "",
) -> Optional[Dict[str, Any]]:
    if df.empty or "Data" not in df.columns:
        return None

    date_key = _normalize_date_key(date_value)
    if not date_key:
        return None

    day_rows = df[df["Data"].apply(_normalize_date_key) == date_key]
    if day_rows.empty:
        return None

    turma_codigo_key = _normalize_text_fold(_normalize_excel_string(requested_turma_codigo))
    turma_label_key = _normalize_text_fold(_normalize_excel_string(requested_turma_label))

    scoped_rows = day_rows
    if turma_codigo_key or turma_label_key:
        turma_codigo_col = day_rows["TurmaCodigo"].astype(str).map(
            lambda value: _normalize_text_fold(_normalize_excel_string(value))
        )
        turma_label_col = day_rows["TurmaLabel"].astype(str).map(
            lambda value: _normalize_text_fold(_normalize_excel_string(value))
        )

        scope_mask = pd.Series([True] * len(day_rows), index=day_rows.index)
        if turma_codigo_key:
            scope_mask = scope_mask & ((turma_codigo_col == turma_codigo_key) | (turma_codigo_col == ""))
        if turma_label_key:
            scope_mask = scope_mask & ((turma_label_col == turma_label_key) | (turma_label_col == ""))

        scoped_rows = day_rows[scope_mask]

    # Compatibilidade retroativa: se não houver escopo claro no histórico,
    # volta para a regra diária para não quebrar leituras antigas.
    if scoped_rows.empty:
        scoped_rows = day_rows

    baseline_row = _select_latest_pool_log_from_rows(scoped_rows, requested_horario)
    if baseline_row is None:
        return None

    # Regra de sincronização: sempre retorna o log mais recente por escopo
    # (dia/turma/horário), sem segmentar por professor/dispositivo.
    return baseline_row

class ImportResult(BaseModel):
    units_created: int
    units_updated: int
    classes_created: int
    classes_updated: int
    students_created: int
    students_updated: int

class ImportStatusOut(BaseModel):
    filename: Optional[str] = None
    last_import_at: Optional[str] = None
    rows_processed: int = 0
    units_created: int = 0
    units_updated: int = 0
    classes_created: int = 0
    classes_updated: int = 0
    students_created: int = 0
    students_updated: int = 0


def _import_status_file() -> str:
    return os.path.join(DATA_DIR, "import_status.json")


def _save_import_status(status: Dict[str, Any]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(_import_status_file(), "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)


def _load_import_status() -> Dict[str, Any]:
    file_path = _import_status_file()
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}

class ImportUnitOut(BaseModel):
    id: int
    name: str

class ImportClassOut(BaseModel):
    id: int
    unit_id: int
    grupo: str
    codigo: str
    turma_label: str
    horario: str
    professor: str
    nivel: str
    faixa_etaria: str
    capacidade: int = Field(ge=0)
    dias_semana: str

class ImportStudentOut(BaseModel):
    id: int
    student_uid: str = ""
    class_id: Optional[int] = None
    nome: str
    whatsapp: str
    data_nascimento: str
    data_atestado: str
    categoria: str
    genero: str
    parq: str
    atestado: bool

class BootstrapOut(BaseModel):
    units: list[ImportUnitOut]
    classes: list[ImportClassOut]
    students: list[ImportStudentOut]

class ImportStudentUpsertPayload(BaseModel):
    nome: str
    student_uid: Optional[str] = ""
    turma: Optional[str] = ""
    horario: Optional[str] = ""
    professor: Optional[str] = ""
    whatsapp: Optional[str] = ""
    data_nascimento: Optional[str] = ""
    data_atestado: Optional[str] = ""
    categoria: Optional[str] = ""
    genero: Optional[str] = ""
    parq: Optional[str] = ""
    atestado: bool = False
    movement_type: str = "correction"


class ImportStudentBulkAllocatePayload(BaseModel):
    student_ids: list[int]
    turma: str
    horario: str
    professor: str
    movement_type: str = "correction"


class MaintenancePurgeMonthPayload(BaseModel):
    month: str = "2026-02"
    clear_transfer_overrides: bool = True
    clear_exclusions: bool = False


class MaintenanceBootstrapResetPayload(BaseModel):
    clear_transfer_overrides: bool = True
    clear_student_uid_registry: bool = True
    clear_import_status: bool = True


class MaintenanceDiagnosticsOut(BaseModel):
    bootstrap: Dict[str, int]
    feb2026: Dict[str, int]
    importStatus: Dict[str, Any]

@app.get("/weather")
def get_weather(date: str):
    cptec_url = os.getenv(
        "CPTEC_WEATHER_URL",
        "http://servicos.cptec.inpe.br/XML/cidade/7dias/5678/previsao.xml",
    )
    tempo_map = {
        "ec": "Encoberto com Chuvas Isoladas",
        "ci": "Chuvas Isoladas",
        "c": "Chuva",
        "in": "Instável",
        "pp": "Possibilidade de Pancadas de Chuva",
        "cm": "Chuva pela Manhã",
        "cn": "Chuva à Noite",
        "pt": "Pancadas de Chuva à Tarde",
        "pm": "Pancadas de Chuva pela Manhã",
        "np": "Nublado e Pancadas de Chuva",
        "pc": "Pancadas de Chuva",
        "pn": "Parcialmente Nublado",
        "cv": "Chuvisco",
        "ch": "Chuvoso",
        "t": "Tempestade",
        "ps": "Predomínio de Sol",
        "sn": "Sol entre Nuvens",
        "e": "Encoberto",
        "n": "Nublado",
        "cl": "Céu Claro",
        "nv": "Nevoeiro",
        "g": "Geada",
        "pnt": "Pancadas de Chuva à Noite",
        "psc": "Possibilidade de Chuva",
        "pcm": "Possibilidade de Chuva pela Manhã",
        "pct": "Possibilidade de Chuva à Tarde",
        "pcn": "Possibilidade de Chuva à Noite",
        "npt": "Nublado com Pancadas à Tarde",
        "npn": "Nublado com Pancadas à Noite",
        "ncn": "Nublado com Possibilidade de Chuva à Noite",
        "nct": "Nublado com Possibilidade de Chuva à Tarde",
        "ncm": "Nublado com Possibilidade de Chuva pela Manhã",
        "npm": "Nublado com Pancadas pela Manhã",
        "npp": "Nublado com Possibilidade de Chuva",
        "vn": "Variação de Nebulosidade",
        "ct": "Chuva à Tarde",
        "ppn": "Possibilidade de Pancadas de Chuva à Noite",
        "ppt": "Possibilidade de Pancadas de Chuva à Tarde",
        "ppm": "Possibilidade de Pancadas de Chuva pela Manhã",
    }

    snapshots = _load_weather_snapshots()
    requested_date = str(date or "").strip()
    snapshot = snapshots.get(requested_date)
    if isinstance(snapshot, dict):
        return {
            "temp": str(snapshot.get("temp") or ""),
            "condition": str(snapshot.get("condition") or ""),
            "conditionCode": str(snapshot.get("conditionCode") or ""),
            "source": "snapshot",
        }

    try:
        resp = requests.get(cptec_url, timeout=10)
        resp.raise_for_status()
        root = ET.fromstring(resp.content)
        previsoes = root.findall(".//previsao")
        if not previsoes:
            return {"temp": "", "condition": "", "conditionCode": "", "source": "unavailable"}

        changed = False
        for item in previsoes:
            dia = (item.findtext("dia") or "").strip()
            if not dia:
                continue
            minima_txt = (item.findtext("minima") or "").strip()
            maxima_txt = (item.findtext("maxima") or "").strip()
            tempo_code = (item.findtext("tempo") or "").strip().lower()
            condition = tempo_map.get(tempo_code, "Parcialmente Nublado")
            temp = _compute_weather_temp(minima_txt, maxima_txt, "26")

            next_snapshot = _build_weather_snapshot(dia, temp, condition, tempo_code)
            prev_snapshot = snapshots.get(dia)
            if not isinstance(prev_snapshot, dict) or any(
                str(prev_snapshot.get(field) or "") != str(next_snapshot.get(field) or "")
                for field in ["temp", "condition", "conditionCode"]
            ):
                snapshots[dia] = next_snapshot
                changed = True

        if changed:
            _save_weather_snapshots(snapshots)

        refreshed = snapshots.get(requested_date)
        if isinstance(refreshed, dict):
            return {
                "temp": str(refreshed.get("temp") or ""),
                "condition": str(refreshed.get("condition") or ""),
                "conditionCode": str(refreshed.get("conditionCode") or ""),
                "source": "snapshot",
            }

        # Para datas passadas sem snapshot salvo, evita devolver previsao do dia atual.
        try:
            today_iso = datetime.utcnow().date().isoformat()
            if requested_date and requested_date < today_iso:
                return {"temp": "", "condition": "", "conditionCode": "", "source": "unavailable"}
        except Exception:
            pass

        first = previsoes[0]
        minima_txt = (first.findtext("minima") or "").strip()
        maxima_txt = (first.findtext("maxima") or "").strip()
        tempo_code = (first.findtext("tempo") or "").strip().lower()
        condition = tempo_map.get(tempo_code, "Parcialmente Nublado")
        temp = _compute_weather_temp(minima_txt, maxima_txt, "26")
        return {
            "temp": str(temp),
            "condition": condition,
            "conditionCode": tempo_code,
            "source": "cptec-fallback",
        }
    except Exception:
        return {"temp": "", "condition": "", "conditionCode": "", "source": "unavailable"}

@app.post("/pool-log")
def append_pool_log(entry: PoolLogEntryModel):
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        file_path = os.path.join(DATA_DIR, "logPiscina.xlsx")
        row = _pool_log_row_from_entry(entry)

        try:
            df = _load_pool_log_dataframe()
        except PermissionError:
            raise HTTPException(status_code=423, detail="logPiscina.xlsx em uso. Feche o arquivo para salvar.")

        latest_for_day = _select_latest_pool_log_for_day(
            df,
            entry.data,
            entry.horario,
            entry.professor,
            entry.turmaCodigo,
            entry.turmaLabel,
        )
        if latest_for_day and _pool_log_meaningful_signature(latest_for_day) == _pool_log_meaningful_signature(row):
            return {"ok": True, "action": "noop", "file": file_path}

        action = "created"
        db_saved = False
        db_error = None

        # Tentar salvar no banco PostgreSQL
        try:
            from app.database import engine as _db_engine
            from sqlmodel import Session as _DBSession

            with _DBSession(_db_engine) as _db:
                cloro_raw = row.get("Cloro (ppm)", None)
                cloro_value = None
                if cloro_raw is not None:
                    try:
                        parsed = float(str(cloro_raw).replace(",", ".").strip())
                        if math.isfinite(parsed):
                            cloro_value = parsed
                    except Exception:
                        cloro_value = None

                _db.add(PoolLog(
                    data=_normalize_date_key(row.get("Data", "")),
                    turma_codigo=str(row.get("TurmaCodigo", "") or "").strip(),
                    turma_label=str(row.get("TurmaLabel", "") or "").strip(),
                    horario=_normalize_horario_key(row.get("Horario", "") or ""),
                    professor=str(row.get("Professor", "") or "").strip(),
                    clima1=str(row.get("Clima 1", "") or "").strip(),
                    clima2=str(row.get("Clima 2", "") or "").strip(),
                    status_aula=str(row.get("Status_aula", "") or "").strip(),
                    nota=str(row.get("Nota", "") or "").strip(),
                    tipo_ocorrencia=str(row.get("Tipo_ocorrencia", "") or "").strip(),
                    temp_externa=str(row.get("Temp. (C)", "") or "").strip(),
                    temp_piscina=str(row.get("Piscina (C)", "") or "").strip(),
                    cloro_ppm=cloro_value,
                    saved_at=pd.Timestamp.utcnow().isoformat(),
                ))
                _db.commit()
                db_saved = True
        except Exception as e:
            db_error = str(e)
            print(f"[WARN] pool-log DB save failed: {db_error}")

        # Tentar salvar no Excel (sempre tenta, mesmo se banco falhar)
        try:
            excel_df = _load_pool_log(file_path)
            excel_df = pd.concat([excel_df, pd.DataFrame([row])], ignore_index=True)
            excel_df.to_excel(file_path, index=False)
        except PermissionError:
            raise HTTPException(status_code=423, detail="logPiscina.xlsx em uso. Feche o arquivo para salvar.")

        # Se banco falhou, tentar avisar ao cliente
        if not db_saved and db_error:
            print(f"[WARN] Pool-log salvo em Excel mas falhou no banco: {db_error}")

        return {"ok": True, "action": action, "file": file_path, "db_saved": db_saved}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"pool-log error: {exc}")

@app.get("/pool-log")
def get_pool_log(
    date: str,
    turmaCodigo: Optional[str] = None,
    turmaLabel: Optional[str] = None,
    horario: Optional[str] = None,
    professor: Optional[str] = None,
):
    try:
        df = _load_pool_log_dataframe()
        if "Data" not in df.columns:
            return Response(status_code=204)

        selected_row = _select_latest_pool_log_for_day(
            df,
            date,
            horario or "",
            professor or "",
            turmaCodigo or "",
            turmaLabel or "",
        )
        if selected_row is None:
            return Response(status_code=204)

        return _pool_log_row_as_response(selected_row)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"pool-log read error: {exc}")

@app.post("/attendance-log")
def append_attendance_log(payload: AttendanceLogPayload):
    try:
        file_path = os.path.join(DATA_DIR, "baseChamada.json")
        item = payload.dict()
        item["horario"] = _normalize_horario_key(item.get("horario") or "")
        item["turmaCodigo"] = str(item.get("turmaCodigo") or "").strip()
        item["turmaLabel"] = str(item.get("turmaLabel") or "").strip()
        item["professor"] = str(item.get("professor") or "").strip()

        # Defensive merge: if a client sends a partial roster snapshot,
        # preserve existing students from the latest log for this class/month.
        latest_logs = _load_latest_attendance_logs(str(item.get("mes") or "").strip() or None)
        latest_same_class: Optional[Dict[str, Any]] = None
        for key in _attendance_log_lookup_keys(item):
            candidate = latest_logs.get(key)
            if candidate:
                latest_same_class = candidate
                break

        # Do not reject by client timestamp: devices can have clock skew.
        # We always merge incoming snapshot with latest server snapshot.

        incoming_registros = item.get("registros") or []
        if latest_same_class and isinstance(latest_same_class.get("registros"), list):
            existing_registros = latest_same_class.get("registros") or []

            def _student_key(value: Any) -> str:
                if not isinstance(value, dict):
                    return ""
                return _normalize_text(str(value.get("aluno_nome") or "").strip())

            merged: Dict[str, Dict[str, Any]] = {}
            for record in existing_registros:
                key = _student_key(record)
                if key:
                    merged[key] = dict(record)

            def _merge_student_record(existing: Dict[str, Any], incoming: Dict[str, Any]) -> Dict[str, Any]:
                existing_attendance = existing.get("attendance") if isinstance(existing.get("attendance"), dict) else {}
                incoming_attendance = incoming.get("attendance") if isinstance(incoming.get("attendance"), dict) else {}

                existing_justifications = existing.get("justifications") if isinstance(existing.get("justifications"), dict) else {}
                incoming_justifications = incoming.get("justifications") if isinstance(incoming.get("justifications"), dict) else {}

                existing_notes = existing.get("notes") if isinstance(existing.get("notes"), list) else []
                incoming_notes_raw = incoming.get("notes")
                incoming_notes = [
                    str(note).strip()
                    for note in incoming_notes_raw
                    if str(note or "").strip()
                ] if isinstance(incoming_notes_raw, list) else None

                normalized_incoming_attendance = {
                    str(date_key).strip(): str(value or "").strip()
                    for date_key, value in incoming_attendance.items()
                    if str(date_key or "").strip()
                }
                normalized_incoming_justifications = {
                    str(date_key).strip(): str(value or "").strip()
                    for date_key, value in incoming_justifications.items()
                    if str(date_key or "").strip()
                }

                merged_attendance = {
                    **existing_attendance,
                }
                for date_key, status_value in normalized_incoming_attendance.items():
                    if status_value:
                        merged_attendance[date_key] = status_value

                non_empty_attendance_dates = {
                    date_key
                    for date_key, status_value in normalized_incoming_attendance.items()
                    if status_value
                }

                merged_justifications = {
                    **existing_justifications,
                }
                for date_key in non_empty_attendance_dates:
                    status_value = str(merged_attendance.get(date_key) or "").strip()
                    incoming_reason = normalized_incoming_justifications.get(date_key, "")
                    if status_value == "Justificado":
                        if incoming_reason:
                            merged_justifications[date_key] = incoming_reason
                    else:
                        merged_justifications.pop(date_key, None)

                for date_key, reason in normalized_incoming_justifications.items():
                    status_value = str(merged_attendance.get(date_key) or "").strip()
                    if reason and status_value == "Justificado":
                        merged_justifications[date_key] = reason

                return {
                    **existing,
                    **incoming,
                    "attendance": merged_attendance,
                    "justifications": merged_justifications,
                    "notes": incoming_notes if incoming_notes is not None else existing_notes,
                }

            for record in incoming_registros:
                key = _student_key(record)
                if key:
                    existing = merged.get(key) or {}
                    merged[key] = _merge_student_record(existing, dict(record))

            item["registros"] = list(merged.values())

        item["saved_at"] = pd.Timestamp.utcnow().isoformat()

        # Gravar no Supabase/PostgreSQL (persistência permanente)
        try:
            from app.database import engine as _db_engine
            from sqlmodel import Session as _DBSession
            _turma_codigo_key = str(item.get("turmaCodigo") or "").strip()
            _horario_key = str(item.get("horario") or "").strip()
            _professor_key = str(item.get("professor") or "").strip()
            _mes_key = str(item.get("mes") or "").strip()
            _incoming_client_mutation_id = item.get("clientMutationId")
            try:
                _incoming_client_mutation_id = int(_incoming_client_mutation_id)
            except Exception:
                _incoming_client_mutation_id = None
            _registros_to_store = item.get("registros") or []
            with _DBSession(_db_engine) as _db:
                _existing = _db.exec(
                    select(AttendanceLog).where(
                        AttendanceLog.turma_codigo == _turma_codigo_key,
                        AttendanceLog.horario == _horario_key,
                        AttendanceLog.professor == _professor_key,
                        AttendanceLog.mes == _mes_key,
                    )
                ).first()
                existing_metadata = _attendance_source_metadata(_existing.source) if _existing else {}
                existing_client_mutation_id = existing_metadata.get("clientMutationId")
                try:
                    existing_client_mutation_id = int(existing_client_mutation_id)
                except Exception:
                    existing_client_mutation_id = None

                if (
                    _existing
                    and _incoming_client_mutation_id is not None
                    and existing_client_mutation_id is not None
                    and _incoming_client_mutation_id < existing_client_mutation_id
                ):
                    return {
                        "ok": True,
                        "skipped": True,
                        "reason": "stale_snapshot",
                        "saved_at": _existing.saved_at,
                    }

                source_metadata: Dict[str, Any] = {}
                if item.get("source"):
                    source_metadata["source"] = item.get("source")
                if _incoming_client_mutation_id is not None:
                    source_metadata["clientMutationId"] = _incoming_client_mutation_id

                serialized_source = json.dumps(source_metadata, ensure_ascii=False) if source_metadata else None
                if _existing:
                    _existing.turma_label = str(item.get("turmaLabel") or "").strip()
                    _existing.saved_at = item["saved_at"]
                    _existing.client_saved_at = str(item.get("clientSavedAt") or "")
                    _existing.source = serialized_source
                    _existing.registros_json = json.dumps(_registros_to_store, ensure_ascii=False)
                    _db.add(_existing)
                else:
                    _log_row = AttendanceLog(
                        turma_codigo=_turma_codigo_key,
                        turma_label=str(item.get("turmaLabel") or "").strip(),
                        horario=_horario_key,
                        professor=_professor_key,
                        mes=_mes_key,
                        saved_at=item["saved_at"],
                        client_saved_at=str(item.get("clientSavedAt") or ""),
                        source=serialized_source,
                        registros_json=json.dumps(_registros_to_store, ensure_ascii=False),
                    )
                    _db.add(_log_row)
                _db.commit()
        except Exception:
            pass  # falha no DB não impede o salvamento em JSON

        _append_json_list(file_path, [item])
        return {"ok": True, "file": file_path}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"attendance-log error: {exc}")

@app.post("/attendance-log/force-sync")
def force_attendance_sync(payload: AttendanceSyncProbePayload):
    try:
        item = payload.dict()
        item["horario"] = _normalize_horario_key(item.get("horario") or "")
        item["turmaCodigo"] = str(item.get("turmaCodigo") or "").strip()
        item["turmaLabel"] = str(item.get("turmaLabel") or "").strip()
        item["professor"] = str(item.get("professor") or "").strip()
        item["mes"] = str(item.get("mes") or "").strip()

        latest_logs = _load_latest_attendance_logs(item.get("mes") or None)
        latest_same_class: Optional[Dict[str, Any]] = None
        for key in _attendance_log_lookup_keys(item):
            candidate = latest_logs.get(key)
            if candidate:
                latest_same_class = candidate
                break

        return {
            "ok": True,
            "hasLog": bool(latest_same_class),
            "saved_at": (latest_same_class or {}).get("saved_at"),
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"attendance-sync error: {exc}")

@app.post("/justifications-log")
def append_justifications_log(entries: List[JustificationLogEntry]):
    try:
        if not entries:
            return {"ok": True, "file": os.path.join(DATA_DIR, "baseJustificativas.json")}

        file_path = os.path.join(DATA_DIR, "baseJustificativas.json")
        existing_items = _load_json_list(file_path)

        def _entry_key(item: Dict[str, Any]) -> str:
            aluno = _normalize_text(str(item.get("aluno_nome") or "").strip())
            data = _normalize_date_key(item.get("data") or "")
            turma = _normalize_text(str(item.get("turmaCodigo") or item.get("turmaLabel") or "").strip())
            horario = _normalize_horario_key(item.get("horario") or "")
            professor = _normalize_text(str(item.get("professor") or "").strip())
            return f"{aluno}||{data}||{turma}||{horario}||{professor}"

        keyed_items: Dict[str, Dict[str, Any]] = {}
        for existing in existing_items:
            if not isinstance(existing, dict):
                continue
            normalized_existing = dict(existing)
            normalized_existing["horario"] = _normalize_horario_key(normalized_existing.get("horario") or "")
            normalized_existing["turmaCodigo"] = str(normalized_existing.get("turmaCodigo") or "").strip()
            normalized_existing["turmaLabel"] = str(normalized_existing.get("turmaLabel") or "").strip()
            normalized_existing["professor"] = str(normalized_existing.get("professor") or "").strip()
            normalized_existing["data"] = _normalize_date_key(normalized_existing.get("data") or "")
            key = _entry_key(normalized_existing)
            if key.strip("|"):
                keyed_items[key] = normalized_existing

        upsert_count = 0
        for entry in entries:
            item = entry.dict()
            item["horario"] = _normalize_horario_key(item.get("horario") or "")
            item["turmaCodigo"] = str(item.get("turmaCodigo") or "").strip()
            item["turmaLabel"] = str(item.get("turmaLabel") or "").strip()
            item["professor"] = str(item.get("professor") or "").strip()
            item["data"] = _normalize_date_key(item.get("data") or "")
            item["saved_at"] = pd.Timestamp.utcnow().isoformat()
            key = _entry_key(item)
            if not key.strip("|"):
                continue
            keyed_items[key] = item
            upsert_count += 1

        _save_json_list(file_path, list(keyed_items.values()))
        return {"ok": True, "file": file_path, "count": upsert_count}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"justifications-log error: {exc}")

def _normalize_text(value: Optional[str]) -> str:
    return _repair_mojibake_text(str(value or "")).strip().lower()


def _text_corruption_score(value: str) -> int:
    if not value:
        return 0

    suspicious_chars = "ÃÂ�├┤┬└┘╜╨║"
    suspicious = sum(1 for ch in value if ch in suspicious_chars)
    box_drawing = sum(1 for ch in value if 0x2500 <= ord(ch) <= 0x257F)
    pipe_between_letters = len(re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ]\|[A-Za-zÀ-ÖØ-öø-ÿ]", value))
    return suspicious + box_drawing + (pipe_between_letters * 3)


def _repair_mojibake_text(value: Optional[str]) -> str:
    raw = str(value or "")
    if not raw:
        return ""

    best = raw
    best_score = _text_corruption_score(raw)

    for encoding in ("cp437", "latin-1", "cp1252"):
        try:
            candidate = raw.encode(encoding).decode("utf-8")
        except Exception:
            continue

        candidate_score = _text_corruption_score(candidate)
        if candidate_score < best_score:
            best = candidate
            best_score = candidate_score

    return best


def _transfer_overrides_file() -> str:
    return os.path.join(DATA_DIR, "studentTransferOverrides.json")


def _student_uid_registry_file() -> str:
    return os.path.join(DATA_DIR, "studentUids.json")


def _normalize_whatsapp_digits(value: Optional[str]) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _student_identity_key(nome: Optional[str], data_nascimento: Optional[str], whatsapp: Optional[str]) -> str:
    name_key = _normalize_text(nome)
    birth_key = str(data_nascimento or "").strip()
    phone_key = _normalize_whatsapp_digits(whatsapp)
    return "|".join([name_key, birth_key, phone_key])


def _student_uid_identity_key(nome: Optional[str], data_nascimento: Optional[str], whatsapp: Optional[str]) -> str:
    name_key = _normalize_text(nome)
    birth_key = str(data_nascimento or "").strip()
    phone_key = _normalize_whatsapp_digits(whatsapp)
    if not name_key:
        return ""
    if not birth_key and not phone_key:
        return ""
    return "|".join([name_key, birth_key, phone_key])


def _load_student_uid_registry() -> Dict[str, str]:
    path = _student_uid_registry_file()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict):
            return {str(k): str(v) for k, v in payload.items() if str(k).strip() and str(v).strip()}
    except Exception:
        pass
    return {}


def _save_student_uid_registry(registry: Dict[str, str]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    path = _student_uid_registry_file()
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(registry, handle, ensure_ascii=False, indent=2)


def _ensure_student_uid_for_student(
    student: models.ImportStudent,
    registry: Optional[Dict[str, str]] = None,
    preferred_uid: Optional[str] = None,
) -> tuple[str, bool]:
    data = registry if registry is not None else _load_student_uid_registry()
    changed = False

    identity_key = _student_uid_identity_key(student.nome, student.data_nascimento, student.whatsapp)
    registry_identity_key = f"identity:{identity_key}" if identity_key else ""
    registry_legacy_key = f"legacy:{int(student.id or 0)}" if student.id else ""
    preferred = str(preferred_uid or "").strip()

    uid = ""
    if preferred:
        uid = preferred
    elif registry_legacy_key and data.get(registry_legacy_key):
        uid = str(data.get(registry_legacy_key) or "")
    elif registry_identity_key and data.get(registry_identity_key):
        uid = str(data.get(registry_identity_key) or "")

    if not uid:
        uid = str(uuid.uuid4())

    if registry_legacy_key and data.get(registry_legacy_key) != uid:
        data[registry_legacy_key] = uid
        changed = True
    if registry_identity_key and data.get(registry_identity_key) != uid:
        data[registry_identity_key] = uid
        changed = True

    if registry is None and changed:
        _save_student_uid_registry(data)

    return uid, changed


def _load_transfer_overrides() -> List[Dict[str, Any]]:
    return _load_json_list(_transfer_overrides_file())


def _save_transfer_overrides(items: List[Dict[str, Any]]) -> None:
    _save_json_list(_transfer_overrides_file(), items)


def _upsert_transfer_override_for_student(student: models.ImportStudent, target_class: models.ImportClass) -> None:
    identity_key = _student_identity_key(student.nome, student.data_nascimento, student.whatsapp)
    if not identity_key:
        return

    items = _load_transfer_overrides()
    payload = {
        "key": identity_key,
        "nome": student.nome,
        "data_nascimento": student.data_nascimento or "",
        "whatsapp": student.whatsapp or "",
        "turmaCodigo": target_class.codigo or "",
        "turmaLabel": target_class.turma_label or target_class.codigo or "",
        "horario": target_class.horario or "",
        "professor": target_class.professor or "",
        "updated_at": datetime.utcnow().isoformat(),
    }

    updated = False
    for idx, item in enumerate(items):
        if str(item.get("key") or "").strip() == identity_key:
            items[idx] = {**item, **payload}
            updated = True
            break

    if not updated:
        items.append(payload)

    _save_transfer_overrides(items)


def _remove_transfer_override_for_student(student: models.ImportStudent) -> None:
    identity_key = _student_identity_key(student.nome, student.data_nascimento, student.whatsapp)
    if not identity_key:
        return

    items = _load_transfer_overrides()
    filtered = [item for item in items if str(item.get("key") or "").strip() != identity_key]
    if len(filtered) != len(items):
        _save_transfer_overrides(filtered)


def _find_class_from_transfer_override(session: Session, override: Dict[str, Any]) -> Optional[models.ImportClass]:
    turma_ref = str(override.get("turmaCodigo") or override.get("turmaLabel") or "").strip()
    horario_ref = str(override.get("horario") or "").strip()
    professor_ref = str(override.get("professor") or "").strip()
    if not turma_ref or not horario_ref or not professor_ref:
        return None
    return _find_import_class_by_triple(session, turma_ref, horario_ref, professor_ref)


def _has_strong_student_identity(student: models.ImportStudent) -> bool:
    return bool(str(student.data_nascimento or "").strip() or _normalize_whatsapp_digits(student.whatsapp))


def _dedupe_import_students(session: Session) -> int:
    students = session.exec(select(models.ImportStudent)).all()
    seen: Dict[tuple[int, str, str], int] = {}
    removed = 0

    for student in students:
        identity = _student_identity_key(student.nome, student.data_nascimento, student.whatsapp)
        key = (student.class_id, _normalize_text(student.nome), identity)
        previous_id = seen.get(key)
        if previous_id is None:
            seen[key] = int(student.id or 0)
            continue

        # Keep the oldest id and remove newer duplicates.
        if int(student.id or 0) < previous_id:
            old = session.get(models.ImportStudent, previous_id)
            if old is not None:
                session.delete(old)
                removed += 1
            seen[key] = int(student.id or 0)
        else:
            session.delete(student)
            removed += 1

    return removed


def _dedupe_import_students_global(session: Session) -> int:
    students = session.exec(select(models.ImportStudent)).all()
    candidates: Dict[str, List[models.ImportStudent]] = {}
    for student in students:
        if not _has_strong_student_identity(student):
            continue
        identity = _student_identity_key(student.nome, student.data_nascimento, student.whatsapp)
        if not identity:
            continue
        candidates.setdefault(identity, []).append(student)

    if not candidates:
        return 0

    removed = 0
    for grouped_students in candidates.values():
        if len(grouped_students) <= 1:
            continue

        # Prefer the allocated student (class_id not None); among equals, keep oldest id.
        keep = min(
            grouped_students,
            key=lambda s: (0 if s.class_id is not None else 1, int(s.id or 0)),
        )
        for student in grouped_students:
            if student.id == keep.id:
                continue
            session.delete(student)
            removed += 1

    return removed


def _apply_transfer_overrides(session: Session) -> int:
    overrides = _load_transfer_overrides()
    if not overrides:
        return 0

    with session.no_autoflush:
        students = session.exec(select(models.ImportStudent)).all()
        classes = session.exec(select(models.ImportClass)).all()

        by_identity: Dict[str, List[models.ImportStudent]] = {}
        for student in students:
            identity = _student_identity_key(student.nome, student.data_nascimento, student.whatsapp)
            if not identity:
                continue
            by_identity.setdefault(identity, []).append(student)

        def _resolve_target_class(override: Dict[str, Any]) -> Optional[models.ImportClass]:
            turma_ref = _normalize_text(override.get("turmaCodigo") or override.get("turmaLabel") or "")
            horario_ref = _normalize_horario_value(str(override.get("horario") or ""))
            professor_ref = _normalize_text(override.get("professor") or "")
            if not turma_ref or not horario_ref or not professor_ref:
                return None

            for cls in classes:
                cls_codigo = _normalize_text(cls.codigo or "")
                cls_label = _normalize_text(cls.turma_label or cls.codigo or "")
                cls_horario = _normalize_horario_value(cls.horario or "")
                cls_professor = _normalize_text(cls.professor or "")
                turma_matches = turma_ref in {cls_codigo, cls_label}
                if turma_matches and cls_horario == horario_ref and cls_professor == professor_ref:
                    return cls
            return None

        moved = 0
        for override in overrides:
            identity = str(override.get("key") or "").strip()
            if not identity:
                identity = _student_identity_key(
                    str(override.get("nome") or ""),
                    str(override.get("data_nascimento") or ""),
                    str(override.get("whatsapp") or ""),
                )
            if not identity:
                continue

            target_class = _resolve_target_class(override)
            if target_class is None:
                continue

            group = list(by_identity.get(identity, []))
            if not group:
                continue

            group.sort(key=lambda item: int(item.id or 0))
            canonical = group[0]

            if canonical.class_id != target_class.id:
                canonical.class_id = target_class.id
                moved += 1

            for duplicate in group[1:]:
                session.delete(duplicate)
                moved += 1

            by_identity[identity] = [canonical]

        moved += _dedupe_import_students(session)
        moved += _dedupe_import_students_global(session)
        return moved


def _to_proper_case(value: Optional[str]) -> str:
    raw = _repair_mojibake_text(value).strip()
    if not raw:
        return ""

    lower_particles = {"da", "de", "do", "das", "dos", "e"}
    tokens = re.split(r"(\s+)", raw)
    words: List[str] = []
    word_index = 0

    for token in tokens:
        if token.isspace() or token == "":
            words.append(token)
            continue

        chunk_parts = re.split(r"([-'])", token)
        rebuilt: List[str] = []
        local_word = ""
        for part in chunk_parts:
            if part in {"-", "'"}:
                rebuilt.append(part)
                continue
            piece = part.strip()
            if not piece:
                rebuilt.append(part)
                continue
            p = piece.lower()
            if word_index > 0 and p in lower_particles:
                local_word = p
            else:
                local_word = p[:1].upper() + p[1:]
            rebuilt.append(local_word)
        words.append("".join(rebuilt))
        word_index += 1

    return "".join(words)


def _normalize_text_fold(value: Optional[str]) -> str:
    raw = _repair_mojibake_text(value).strip().lower()
    if not raw:
        return ""
    return "".join(ch for ch in unicodedata.normalize("NFD", raw) if unicodedata.category(ch) != "Mn")


def _infer_schedule_group(turma_label: str, turma_codigo: str) -> str:
    label = _normalize_text_fold(turma_label)
    code = _normalize_text_fold(turma_codigo)
    if ("terca" in label and "quinta" in label) or "tq" in code:
        return "tq"
    if ("quarta" in label and "sexta" in label) or "qs" in code:
        return "qs"
    return "other"


def _load_allowed_schedule_days(today: date) -> Dict[str, set[str]]:
    allowed = {"tq": set(), "qs": set()}
    calendar_path = os.path.join(DATA_DIR, "academicCalendar.json")
    if not os.path.exists(calendar_path):
        return allowed

    try:
        with open(calendar_path, "r", encoding="utf-8") as handle:
            calendar_payload = json.load(handle)
    except Exception:
        return allowed

    settings = calendar_payload.get("settings") or {}
    events = calendar_payload.get("events") or []

    start_raw = str(settings.get("inicioAulas") or "").strip()
    if not start_raw:
        return allowed

    try:
        start_date = datetime.strptime(start_raw, "%Y-%m-%d").date()
    except Exception:
        return allowed

    closed_days = set()
    for event in events:
        day = str((event or {}).get("date") or "").strip()
        if day:
            closed_days.add(day)

    cursor = start_date
    while cursor <= today:
        key = cursor.isoformat()
        if key in closed_days:
            cursor += timedelta(days=1)
            continue
        weekday = cursor.weekday()  # mon=0 ... sun=6
        if weekday in {1, 3}:  # terça/quinta
            allowed["tq"].add(key)
        if weekday in {2, 4}:  # quarta/sexta
            allowed["qs"].add(key)
        cursor += timedelta(days=1)

    return allowed

def _normalize_exclusion_item(item: Dict[str, Any]) -> Dict[str, Any]:
    normalized = dict(item or {})

    for key in ["id", "student_uid", "studentUid", "nome", "Nome", "turma", "Turma", "turmaLabel", "TurmaLabel", "turmaCodigo", "TurmaCodigo", "professor", "Professor", "dataExclusao", "DataExclusao", "motivo_exclusao", "MotivoExclusao"]:
        if key in normalized and normalized.get(key) is not None:
            normalized[key] = str(normalized.get(key)).strip()

    if normalized.get("horario") is not None:
        normalized["horario"] = _normalize_horario_key(normalized.get("horario"))
    if normalized.get("Horario") is not None:
        normalized["Horario"] = _normalize_horario_key(normalized.get("Horario"))

    return normalized


def _exclusion_turma_set(item: Dict[str, Any]) -> set[str]:
    values = {
        _normalize_text(item.get("turma") or item.get("Turma") or ""),
        _normalize_text(item.get("turmaLabel") or item.get("TurmaLabel") or ""),
        _normalize_text(item.get("turmaCodigo") or item.get("TurmaCodigo") or ""),
    }
    return {value for value in values if value}


def _exclusion_records_match(item: Dict[str, Any], payload_dict: Dict[str, Any]) -> bool:
    """Match exclusion records prioritizing context to avoid homonímia collisions."""
    # High-confidence matches (UID or ID)
    item_uid = str(item.get("student_uid") or item.get("studentUid") or "").strip()
    payload_uid = str(payload_dict.get("student_uid") or payload_dict.get("studentUid") or "").strip()
    if item_uid and payload_uid and item_uid == payload_uid:
        return True

    item_id = str(item.get("id") or "").strip()
    payload_id = str(payload_dict.get("id") or "").strip()
    if item_id and payload_id and item_id == payload_id:
        return True

    # Low-confidence match (name only): ALWAYS require context to avoid homonímia
    item_nome = _normalize_text(item.get("nome") or item.get("Nome") or "")
    payload_nome = _normalize_text(payload_dict.get("nome") or payload_dict.get("Nome") or "")
    if not item_nome or not payload_nome or item_nome != payload_nome:
        return False  # Names must match as baseline

    # Now check context FIRST before considering name-match valid
    item_turmas = _exclusion_turma_set(item)
    payload_turmas = _exclusion_turma_set(payload_dict)
    has_turma_context = bool(item_turmas) and bool(payload_turmas)
    turma_matches = not has_turma_context or bool(item_turmas.intersection(payload_turmas))

    item_horario = _normalize_horario_key(item.get("horario") or item.get("Horario") or "")
    payload_horario = _normalize_horario_key(payload_dict.get("horario") or payload_dict.get("Horario") or "")
    has_horario_context = bool(item_horario) and bool(payload_horario)
    horario_matches = not has_horario_context or item_horario == payload_horario

    item_professor = _normalize_text(item.get("professor") or item.get("Professor") or "")
    payload_professor = _normalize_text(payload_dict.get("professor") or payload_dict.get("Professor") or "")
    has_professor_context = bool(item_professor) and bool(payload_professor)
    professor_matches = not has_professor_context or item_professor == payload_professor

    # Accept match only if all provided context matches (no mismatches allowed)
    context_valid = (not has_turma_context or turma_matches) and \
                    (not has_horario_context or horario_matches) and \
                    (not has_professor_context or professor_matches)

    # Require at least ONE context field to be present (avoid bare name matching)
    has_any_context = has_turma_context or has_horario_context or has_professor_context
    return context_valid and has_any_context


def _clean_exclusions_list(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Deduplicate and validate exclusion records. Logs discarded items."""
    cleaned: List[Dict[str, Any]] = []
    discarded_count = 0
    
    for raw in items or []:
        if not isinstance(raw, dict):
            continue

        item = _normalize_exclusion_item(raw)
        uid = str(item.get("student_uid") or item.get("studentUid") or "").strip()
        item_id = str(item.get("id") or "").strip()
        nome = _normalize_text(item.get("nome") or item.get("Nome") or "")
        
        # Validate: must have at least one identifier
        if not uid and not item_id and not nome:
            discarded_count += 1
            continue

        # Dedup using strict matching
        existing_idx = next(
            (idx for idx, existing in enumerate(cleaned) if _exclusion_records_match(existing, item)),
            -1,
        )
        if existing_idx >= 0:
            # Merge: preserve existing + override with new values
            cleaned[existing_idx] = {**cleaned[existing_idx], **item}
        else:
            cleaned.append(item)
    
    if discarded_count > 0:
        import logging
        logging.warning(f"_clean_exclusions_list: discarded {discarded_count} items (missing all identifiers)")

    return cleaned


def _exclusions_file_path() -> str:
    return os.path.join(DATA_DIR, "excludedStudents.json")


def _load_exclusions_from_db() -> List[Dict[str, Any]]:
    with Session(engine) as db:
        rows = db.exec(select(ExclusionRecord).order_by(ExclusionRecord.id.asc())).all()
        items: List[Dict[str, Any]] = []
        for row in rows:
            payload: Dict[str, Any] = {}
            try:
                parsed = json.loads(row.payload_json or "{}")
                if isinstance(parsed, dict):
                    payload = parsed
            except Exception:
                payload = {}

            item: Dict[str, Any] = {**payload}
            if row.exclusion_id and not item.get("id"):
                item["id"] = row.exclusion_id
            if row.student_uid and not item.get("student_uid"):
                item["student_uid"] = row.student_uid
            if row.nome and not item.get("nome"):
                item["nome"] = row.nome
            if row.turma and not item.get("turma"):
                item["turma"] = row.turma
            if row.turma_codigo and not item.get("turmaCodigo"):
                item["turmaCodigo"] = row.turma_codigo
            if row.horario and not item.get("horario"):
                item["horario"] = row.horario
            if row.professor and not item.get("professor"):
                item["professor"] = row.professor
            if row.data_exclusao and not item.get("dataExclusao"):
                item["dataExclusao"] = row.data_exclusao
            if row.motivo_exclusao and not item.get("motivo_exclusao"):
                item["motivo_exclusao"] = row.motivo_exclusao
            items.append(item)
        return items


def _save_exclusions_to_db(items: List[Dict[str, Any]]) -> None:
    normalized_items = [_normalize_exclusion_item(item) for item in (items or []) if isinstance(item, dict)]
    with Session(engine) as db:
        existing = db.exec(select(ExclusionRecord)).all()
        for row in existing:
            db.delete(row)

        now_iso = datetime.utcnow().isoformat()
        for item in normalized_items:
            record = ExclusionRecord(
                exclusion_id=str(item.get("id") or "").strip(),
                student_uid=str(item.get("student_uid") or item.get("studentUid") or "").strip(),
                nome=str(item.get("nome") or item.get("Nome") or "").strip(),
                turma=str(item.get("turma") or item.get("Turma") or item.get("turmaLabel") or item.get("TurmaLabel") or "").strip(),
                turma_codigo=str(item.get("turmaCodigo") or item.get("TurmaCodigo") or item.get("grupo") or item.get("Grupo") or "").strip(),
                horario=_normalize_horario_key(item.get("horario") or item.get("Horario") or ""),
                professor=str(item.get("professor") or item.get("Professor") or "").strip(),
                data_exclusao=str(item.get("dataExclusao") or item.get("DataExclusao") or "").strip(),
                motivo_exclusao=str(item.get("motivo_exclusao") or item.get("MotivoExclusao") or "").strip(),
                payload_json=json.dumps(item, ensure_ascii=False),
                saved_at=str(item.get("saved_at") or now_iso),
            )
            db.add(record)

        db.commit()


def _read_exclusions_state(clean: bool = True) -> List[Dict[str, Any]]:
    file_items = _load_json_list(_exclusions_file_path())
    if file_items:
        items = _clean_exclusions_list(file_items) if clean else file_items
        try:
            _save_exclusions_to_db(items)
        except Exception:
            pass
        return items

    db_items: List[Dict[str, Any]] = []
    try:
        db_items = _load_exclusions_from_db()
    except Exception:
        db_items = []

    if db_items:
        items = _clean_exclusions_list(db_items) if clean else db_items
        try:
            _save_json_list(_exclusions_file_path(), items)
        except Exception:
            pass
        return items

    return []


def _write_exclusions_state(items: List[Dict[str, Any]], clean: bool = True) -> List[Dict[str, Any]]:
    """Write exclusions to DB and file with optional cleaning. Always normalizes."""
    if clean:
        # Cleanliness-first: deduplicate before saving
        payload = _clean_exclusions_list(items or [])
    else:
        # Just normalize, no dedup
        payload = [
            _normalize_exclusion_item(item) for item in (items or []) if isinstance(item, dict)
        ]
    
    # Save to both DB and file atomically where possible
    _save_exclusions_to_db(payload)
    try:
        _save_json_list(_exclusions_file_path(), payload)
    except Exception as e:
        import logging
        logging.error(f"Failed to save exclusions to JSON file: {e}")
    
    return payload


def _list_exclusions_backups(limit: int = 30) -> List[Dict[str, Any]]:
    archive_dir = os.path.join(DATA_DIR, "archive")
    if not os.path.isdir(archive_dir):
        return []

    candidates: List[Dict[str, Any]] = []
    for file_name in os.listdir(archive_dir):
        if not file_name.startswith("excludedStudents_") or not file_name.endswith(".json"):
            continue
        full_path = os.path.join(archive_dir, file_name)
        if not os.path.isfile(full_path):
            continue
        items = _load_json_list(full_path)
        candidates.append(
            {
                "file": file_name,
                "count": len(items),
                "path": full_path,
                "modified_at": datetime.fromtimestamp(os.path.getmtime(full_path)).isoformat(),
            }
        )

    candidates.sort(key=lambda item: str(item.get("modified_at") or ""), reverse=True)
    return candidates[: max(1, int(limit or 30))]


def _merge_exclusions(base_items: List[Dict[str, Any]], incoming_items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Merge exclusion lists efficiently: clean once upfront, merge, then finalize clean."""
    # Clean both inputs once
    base_cleaned = _clean_exclusions_list(base_items or [])
    incoming_cleaned = _clean_exclusions_list(incoming_items or [])
    
    # Merge incoming into base
    merged = list(base_cleaned)  # Shallow copy
    for incoming in incoming_cleaned:
        idx = next(
            (existing_idx for existing_idx, existing in enumerate(merged) if _exclusion_records_match(existing, incoming)),
            -1,
        )
        if idx >= 0:
            merged[idx] = {**merged[idx], **incoming}  # Merge fields
        else:
            merged.append(incoming)  # Add new record
    
    # Final clean pass to catch any dedup edge cases introduced during merge
    return _clean_exclusions_list(merged)

def _resolve_exclusion_match(item: Dict[str, Any], payload: ExclusionEntry) -> bool:
    return _exclusion_records_match(_normalize_exclusion_item(item), _normalize_exclusion_item(payload.dict()))

def _normalize_horario_key(value: Optional[str]) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 3:
        return f"0{digits}"
    if len(digits) >= 4:
        return digits[:4]
    return digits


def _saved_at_sort_key(value: Any) -> int:
    raw = str(value or "").strip()
    if not raw:
        return -1
    try:
        parsed = pd.to_datetime(raw, errors="coerce", utc=True, dayfirst=True)
        if pd.isna(parsed):
            return -1
        return int(parsed.value)
    except Exception:
        return -1


def _attendance_source_metadata(source: Any) -> Dict[str, Any]:
    raw = str(source or "").strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {"source": raw}
    except Exception:
        return {"source": raw}


def _extract_month_key(value: Any) -> str:
    iso = _normalize_date_key(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", iso):
        return iso[:7]
    return ""


def _entry_contains_month(entry: Dict[str, Any], month: str) -> bool:
    if str(entry.get("mes") or "").strip() == month:
        return True

    for key in ["data", "dataExclusao", "saved_at"]:
        if _extract_month_key(entry.get(key)) == month:
            return True

    registros = entry.get("registros") or []
    if isinstance(registros, list):
        for record in registros:
            attendance = (record or {}).get("attendance") or {}
            if isinstance(attendance, dict):
                for date_key in attendance.keys():
                    if _extract_month_key(date_key) == month:
                        return True

            justifications = (record or {}).get("justifications") or {}
            if isinstance(justifications, dict):
                for date_key in justifications.keys():
                    if _extract_month_key(date_key) == month:
                        return True

    return False


def _purge_month_from_json_file(file_path: str, month: str) -> Dict[str, int]:
    items = _load_json_list(file_path)
    if not items:
        return {"before": 0, "after": 0, "removed": 0}

    kept = [item for item in items if not _entry_contains_month(item, month)]
    removed = len(items) - len(kept)
    if removed > 0:
        _save_json_list(file_path, kept)
    return {"before": len(items), "after": len(kept), "removed": removed}


def _purge_month_from_pool_log(month: str) -> Dict[str, int]:
    file_path = os.path.join(DATA_DIR, "logPiscina.xlsx")
    if not os.path.exists(file_path):
        return {"before": 0, "after": 0, "removed": 0}

    df = _load_pool_log(file_path)
    before = len(df)
    if before == 0:
        return {"before": 0, "after": 0, "removed": 0}

    keep_mask = ~df["Data"].apply(lambda value: _extract_month_key(value) == month)
    next_df = df[keep_mask].copy()
    removed = before - len(next_df)
    if removed > 0:
        try:
            next_df.to_excel(file_path, index=False)
        except PermissionError:
            raise HTTPException(status_code=423, detail="logPiscina.xlsx em uso. Feche o arquivo para limpar fevereiro.")

    return {"before": before, "after": len(next_df), "removed": removed}


def _count_month_entries_in_json(file_path: str, month: str) -> int:
    items = _load_json_list(file_path)
    if not items:
        return 0
    return sum(1 for item in items if _entry_contains_month(item, month))


def _count_month_entries_in_pool_log(month: str) -> int:
    file_path = os.path.join(DATA_DIR, "logPiscina.xlsx")
    if not os.path.exists(file_path):
        return 0
    df = _load_pool_log(file_path)
    if len(df) == 0:
        return 0
    return int(df["Data"].apply(lambda value: _extract_month_key(value) == month).sum())


def _purge_month_data(
    month: str,
    clear_transfer_overrides: bool = True,
    clear_exclusions: bool = False,
) -> Dict[str, Any]:
    if not re.fullmatch(r"\d{4}-\d{2}", str(month or "").strip()):
        raise HTTPException(status_code=400, detail="month must be in YYYY-MM format")

    month = str(month).strip()
    os.makedirs(DATA_DIR, exist_ok=True)

    chamada_stats = _purge_month_from_json_file(os.path.join(DATA_DIR, "baseChamada.json"), month)
    justificativa_stats = _purge_month_from_json_file(os.path.join(DATA_DIR, "baseJustificativas.json"), month)
    if clear_exclusions:
        exclusao_stats = _purge_month_from_json_file(os.path.join(DATA_DIR, "excludedStudents.json"), month)
    else:
        total_exclusions = len(_load_json_list(os.path.join(DATA_DIR, "excludedStudents.json")))
        month_exclusions = _count_month_entries_in_json(os.path.join(DATA_DIR, "excludedStudents.json"), month)
        exclusao_stats = {
            "before": total_exclusions,
            "after": total_exclusions,
            "removed": 0,
            "skipped": True,
            "month_matches": month_exclusions,
        }

    snapshots_path = _weather_snapshots_file()
    snapshots = _load_weather_snapshots()
    snapshots_before = len(snapshots)
    snapshots_kept = {k: v for k, v in snapshots.items() if not str(k).startswith(f"{month}-")}
    snapshots_removed = snapshots_before - len(snapshots_kept)
    if snapshots_removed > 0:
        _save_weather_snapshots(snapshots_kept)

    pool_stats = _purge_month_from_pool_log(month)

    overrides_cleared = False
    overrides_count = 0
    if clear_transfer_overrides:
        overrides = _load_transfer_overrides()
        overrides_count = len(overrides)
        if overrides_count > 0:
            _save_transfer_overrides([])
            overrides_cleared = True

    status_cleared = False
    status_path = _import_status_file()
    if os.path.exists(status_path):
        try:
            os.remove(status_path)
            status_cleared = True
        except Exception:
            status_cleared = False

    return {
        "ok": True,
        "month": month,
        "attendance": chamada_stats,
        "justifications": justificativa_stats,
        "exclusions": exclusao_stats,
        "weatherSnapshots": {
            "before": snapshots_before,
            "after": len(snapshots_kept),
            "removed": snapshots_removed,
        },
        "poolLog": pool_stats,
        "transferOverrides": {
            "cleared": overrides_cleared,
            "removed": overrides_count if overrides_cleared else 0,
        },
        "importStatus": {
            "cleared": status_cleared,
        },
    }


@app.post("/maintenance/clear-transfer-overrides")
def clear_transfer_overrides():
    overrides = _load_transfer_overrides()
    removed = len(overrides)
    if removed > 0:
        _save_transfer_overrides([])
    return {
        "ok": True,
        "removed": removed,
    }


@app.post("/maintenance/purge-month-data")
def purge_month_data(payload: MaintenancePurgeMonthPayload):
    return _purge_month_data(
        month=payload.month,
        clear_transfer_overrides=payload.clear_transfer_overrides,
        clear_exclusions=payload.clear_exclusions,
    )


@app.post("/maintenance/reset-bootstrap-data")
def reset_bootstrap_data(
    payload: MaintenanceBootstrapResetPayload,
    session: Session = Depends(get_session),
):
    students = session.exec(select(models.ImportStudent)).all()
    classes = session.exec(select(models.ImportClass)).all()
    units = session.exec(select(models.ImportUnit)).all()

    removed_students = len(students)
    removed_classes = len(classes)
    removed_units = len(units)

    for student in students:
        session.delete(student)
    for cls in classes:
        session.delete(cls)
    for unit in units:
        session.delete(unit)

    session.commit()

    transfer_removed = 0
    transfer_cleared = False
    if payload.clear_transfer_overrides:
        overrides = _load_transfer_overrides()
        transfer_removed = len(overrides)
        if transfer_removed > 0:
            _save_transfer_overrides([])
            transfer_cleared = True

    uid_removed = False
    if payload.clear_student_uid_registry:
        uid_path = _student_uid_registry_file()
        if os.path.exists(uid_path):
            try:
                os.remove(uid_path)
                uid_removed = True
            except Exception:
                uid_removed = False

    status_cleared = False
    if payload.clear_import_status:
        status_path = _import_status_file()
        if os.path.exists(status_path):
            try:
                os.remove(status_path)
                status_cleared = True
            except Exception:
                status_cleared = False

    return {
        "ok": True,
        "removed": {
            "students": removed_students,
            "classes": removed_classes,
            "units": removed_units,
        },
        "transferOverrides": {
            "cleared": transfer_cleared,
            "removed": transfer_removed,
        },
        "studentUidRegistry": {
            "cleared": uid_removed,
        },
        "importStatus": {
            "cleared": status_cleared,
        },
    }


@app.get("/maintenance/diagnostics", response_model=MaintenanceDiagnosticsOut)
def get_maintenance_diagnostics(session: Session = Depends(get_session)):
    month = "2026-02"
    units_count = len(session.exec(select(models.ImportUnit)).all())
    classes_count = len(session.exec(select(models.ImportClass)).all())
    students_count = len(session.exec(select(models.ImportStudent)).all())

    feb_attendance = _count_month_entries_in_json(os.path.join(DATA_DIR, "baseChamada.json"), month)
    feb_justifications = _count_month_entries_in_json(os.path.join(DATA_DIR, "baseJustificativas.json"), month)
    feb_exclusions = _count_month_entries_in_json(os.path.join(DATA_DIR, "excludedStudents.json"), month)
    snapshots = _load_weather_snapshots()
    feb_snapshots = sum(1 for key in snapshots.keys() if str(key).startswith(f"{month}-"))
    feb_pool = _count_month_entries_in_pool_log(month)

    return MaintenanceDiagnosticsOut(
        bootstrap={
            "units": units_count,
            "classes": classes_count,
            "students": students_count,
        },
        feb2026={
            "attendance": feb_attendance,
            "justifications": feb_justifications,
            "exclusions": feb_exclusions,
            "weatherSnapshots": feb_snapshots,
            "poolLog": feb_pool,
        },
        importStatus=_load_import_status(),
    )

def _exclusion_matches_class(entry: Dict[str, Any], cls: models.ImportClass) -> bool:
    cls_codigo = _normalize_text(cls.codigo or "")
    cls_label = _normalize_text(cls.turma_label or cls.codigo or "")
    cls_horario = _normalize_horario_key(cls.horario or "")
    cls_professor = _normalize_text(cls.professor or "")

    ex_codigo = _normalize_text(entry.get("turmaCodigo") or entry.get("TurmaCodigo") or "")
    ex_label = _normalize_text(
        entry.get("turmaLabel")
        or entry.get("TurmaLabel")
        or entry.get("turma")
        or entry.get("Turma")
        or ""
    )
    ex_horario = _normalize_horario_key(entry.get("horario") or entry.get("Horario") or "")
    ex_professor = _normalize_text(entry.get("professor") or entry.get("Professor") or "")

    if ex_codigo and ex_codigo not in {cls_codigo, cls_label}:
        return False
    if ex_label and ex_label not in {cls_label, cls_codigo}:
        return False
    if ex_horario and cls_horario and ex_horario != cls_horario:
        return False
    if ex_professor and cls_professor and ex_professor != cls_professor:
        return False
    return True


def _build_report_student_identity(
    student: models.ImportStudent,
    cls: models.ImportClass,
    student_uid: str = "",
) -> Dict[str, Any]:
    turma_codigo = str(cls.codigo or "").strip()
    turma_label = str(cls.turma_label or cls.codigo or "").strip()
    return {
        "id": str(student.id or "").strip(),
        "student_uid": str(student_uid or "").strip(),
        "nome": str(student.nome or "").strip(),
        "turma": turma_label,
        "turmaLabel": turma_label,
        "turmaCodigo": turma_codigo,
        "grupo": turma_codigo,
        "horario": str(cls.horario or "").strip(),
        "professor": str(cls.professor or "").strip(),
    }


def _is_student_excluded_for_report(
    student: models.ImportStudent,
    cls: models.ImportClass,
    exclusions: List[Dict[str, Any]],
    student_uid: str = "",
) -> bool:
    student_payload = _build_report_student_identity(student, cls, student_uid=student_uid)
    return any(
        _exclusion_matches_class(entry, cls)
        and _exclusion_records_match(entry, student_payload)
        for entry in exclusions
    )

def _map_attendance_value(value: str) -> str:
    normalized = _normalize_text(value)
    if normalized == "presente":
        return "c"
    if normalized == "falta":
        return "f"
    if normalized == "justificado":
        return "j"
    return ""

def _report_day_key(date_value: str) -> str:
    if not date_value:
        return ""
    parts = str(date_value).split("-")
    if len(parts) >= 3:
        return parts[2].zfill(2)
    return str(date_value).strip()

def _sort_report_days(days: List[str]) -> List[str]:
    unique_days = {str(day).strip() for day in days if str(day).strip()}

    def _day_sort_key(day: str):
        if day.isdigit():
            return (0, int(day))
        return (1, day)

    return sorted(unique_days, key=_day_sort_key)

def _resolve_export_targets(payload: ExcelExportPayload, reports: List[ReportClass]) -> List[ReportClass]:
    requested: List[ExportClassSelection] = []
    if payload.classes:
        requested = payload.classes
    elif payload.turma and payload.horario and payload.professor:
        requested = [
            ExportClassSelection(
                turma=payload.turma,
                horario=payload.horario,
                professor=payload.professor,
            )
        ]

    if not requested:
        raise HTTPException(status_code=400, detail="No class selection informed")

    selected: List[ReportClass] = []
    seen_keys = set()
    for req in requested:
        code_key = _normalize_text(req.turmaCodigo)
        turma_key = _normalize_text(req.turma)
        horario_key = _normalize_text(req.horario)
        professor_key = _normalize_text(req.professor)
        key = (code_key, turma_key, horario_key, professor_key)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        found = None
        if code_key:
            exact_code_matches = [
                cls
                for cls in reports
                if _normalize_text(cls.turmaCodigo) == code_key
                or _normalize_text(getattr(cls, "codigo", "")) == code_key
            ]
            if exact_code_matches:
                found = exact_code_matches[0]
        if found is None:
            exact_matches = [
                cls
                for cls in reports
                if _normalize_text(cls.turma) == turma_key
                and _normalize_text(cls.horario) == horario_key
                and _normalize_text(cls.professor) == professor_key
            ]
            if exact_matches:
                found = exact_matches[0]
        if found is None:
            partial_matches = [
                cls
                for cls in reports
                if _normalize_text(cls.turma) == turma_key
                and _normalize_text(cls.horario) == horario_key
            ]
            if len(partial_matches) == 1:
                found = partial_matches[0]
            elif partial_matches:
                if professor_key:
                    found = next(
                        (
                            cls
                            for cls in partial_matches
                            if professor_key in _normalize_text(cls.professor)
                            or _normalize_text(cls.professor) in professor_key
                        ),
                        partial_matches[0],
                    )
                else:
                    found = partial_matches[0]
            else:
                found = next(
                    (
                        cls
                        for cls in reports
                        if _normalize_text(cls.turma) == turma_key
                        and (not professor_key or _normalize_text(cls.professor) == professor_key)
                    ),
                    None,
                )
        if found:
            selected.append(found)

    if not selected:
        raise HTTPException(status_code=404, detail="Class report not found for selected filters")
    return selected

def _build_sheet_title(selected: ReportClass, existing_titles: set[str]) -> str:
    start = _format_horario(selected.horario or "")
    base = f"{start}|{selected.turma}"
    try:
        start_dt = datetime.strptime(start, "%H:%M")
        end_dt = start_dt + timedelta(minutes=45)
        base = f"{start_dt.strftime('%H:%M')}|{end_dt.strftime('%H:%M')}"
    except Exception:
        pass

    safe = re.sub(r"[\[\]:*?/\\]", "-", base).strip() or "Turma"
    safe = safe[:31]
    candidate = safe
    suffix = 1
    while candidate in existing_titles:
        suffix += 1
        label_suffix = f"_{suffix}"
        candidate = f"{safe[: max(1, 31 - len(label_suffix))]}{label_suffix}"
    return candidate

def _populate_attendance_sheet(ws, selected: ReportClass, month: Optional[str], session: Session) -> int:
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    ws["A1"] = "Modalidade:"
    ws["B1"] = "Natação"
    ws["D1"] = "PREFEITURA MUNICIPAL DE VINHEDO"

    ws["A2"] = "Local:"
    ws["B2"] = "Piscina Bela Vista"
    ws["D2"] = "SECRETARIA DE ESPORTE E LAZER"

    ws["A3"] = "Professor:"
    ws["B3"] = selected.professor

    ws["A4"] = "Turma:"
    ws["B4"] = selected.turma
    ws["D4"] = "Nível:"
    ws["E4"] = selected.nivel

    ws["A5"] = "Horário:"
    ws["B5"] = _format_horario(selected.horario or "")
    ws["D5"] = "Mês:"
    ws["E5"] = _format_month_label(month)

    for cell_ref in ("A1", "A2", "A3", "A4", "A5", "D1", "D2", "D4", "D5"):
        ws[cell_ref].font = header_font
        ws[cell_ref].alignment = left_alignment
    for cell_ref in ("B1", "B2", "B3", "B4", "B5", "E4", "E5"):
        ws[cell_ref].alignment = left_alignment

    class_days = _sort_report_days([
        day for aluno in selected.alunos for day in (aluno.historico or {}).keys()
    ])
    header_row = 6
    data_start_row = 7
    date_col_start = 5

    notes_col = None
    for col in range(date_col_start, ws.max_column + 1):
        header_value = str(ws.cell(row=header_row, column=col).value or "").strip()
        header_norm = _normalize_text(header_value)
        if header_norm in {"observacoes", "anotacoes"}:
            notes_col = col
            break

    if notes_col is None:
        notes_col = date_col_start + max(1, len(class_days))

    date_columns = list(range(date_col_start, notes_col))
    visible_days = class_days[:len(date_columns)]

    for col in date_columns:
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    ws.cell(row=header_row, column=1, value="Nome")
    ws.cell(row=header_row, column=2, value="Whatsapp")
    ws.cell(row=header_row, column=3, value="parQ")
    ws.cell(row=header_row, column=4, value="Aniversário")

    for col in range(1, 5):
        ws.cell(row=header_row, column=col).font = header_font
        ws.cell(row=header_row, column=col).alignment = center_alignment

    for idx, col in enumerate(date_columns):
        _copy_cell_style(ws, header_row, date_col_start, header_row, col)
        if idx < len(visible_days):
            day_raw = str(visible_days[idx]).strip()
            day_value: Any = int(day_raw) if day_raw.isdigit() else day_raw
            ws.cell(row=header_row, column=col, value=day_value)
        else:
            ws.cell(row=header_row, column=col, value="")
        ws.cell(row=header_row, column=col).font = header_font
        ws.cell(row=header_row, column=col).alignment = center_alignment

    _copy_cell_style(ws, header_row, date_col_start, header_row, notes_col)
    ws.cell(row=header_row, column=notes_col, value="Anotações")
    ws.cell(row=header_row, column=notes_col).font = header_font
    ws.cell(row=header_row, column=notes_col).alignment = center_alignment

    if ws.max_column > notes_col:
        ws.delete_cols(notes_col + 1, ws.max_column - notes_col)

    details_map = _build_import_student_details_map(
        session=session,
        turma=selected.turma,
        horario=selected.horario,
        professor=selected.professor,
    )

    students_sorted = sorted(selected.alunos, key=lambda item: item.nome)
    full_cell_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for idx, aluno in enumerate(students_sorted):
        row = data_start_row + idx
        for col in range(1, notes_col + 1):
            _copy_cell_style(ws, data_start_row, min(col, date_col_start), row, col)
            ws.cell(row=row, column=col).border = full_cell_border

        student_meta = details_map.get(_normalize_text(aluno.nome), {})
        ws.cell(row=row, column=1, value=aluno.nome)
        ws.cell(row=row, column=2, value=student_meta.get("whatsapp") or "")
        if student_meta.get("atestado"):
            ws.cell(row=row, column=3, value=student_meta.get("data_atestado") or "Com Atestado")
        else:
            ws.cell(row=row, column=3, value=student_meta.get("parq") or "")
        ws.cell(row=row, column=4, value=student_meta.get("data_nascimento") or "")

        for day_idx, day in enumerate(visible_days):
            value = (aluno.historico or {}).get(day, "")
            ws.cell(row=row, column=date_col_start + day_idx, value=value)

        for col in date_columns[len(visible_days):]:
            ws.cell(row=row, column=col, value="")

        ws.cell(row=row, column=notes_col, value=aluno.anotacoes or "")

    name_col_letter = get_column_letter(1)
    notes_col_letter = get_column_letter(notes_col)
    max_name_len = max([len("Nome"), *[len(str(aluno.nome or "")) for aluno in students_sorted]]) if students_sorted else len("Nome")
    ws.column_dimensions[name_col_letter].width = min(50, max(18, max_name_len + 2))
    current_notes_width = ws.column_dimensions[notes_col_letter].width or 0
    ws.column_dimensions[notes_col_letter].width = max(40, current_notes_width)

    end_row = data_start_row + len(students_sorted) - 1
    if end_row < header_row:
        end_row = header_row

    for row in range(header_row, end_row + 1):
        for col in range(1, notes_col + 1):
            horizontal = "left" if col in (1, notes_col) else "center"
            ws.cell(row=row, column=col).alignment = Alignment(horizontal=horizontal, vertical="center")

    return len(students_sorted)

def _build_excel_export_workbook(selected_reports: List[ReportClass], month: Optional[str], session: Session) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    existing_titles: set[str] = set()
    for selected in selected_reports:
        ws = workbook.create_sheet()
        _populate_attendance_sheet(ws=ws, selected=selected, month=month, session=session)
        sheet_title = _build_sheet_title(selected, existing_titles)
        ws.title = sheet_title
        existing_titles.add(sheet_title)

    return workbook

def _build_chamada_pdf(selected_reports: List[ReportClass], month: Optional[str], session: Session) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF export unavailable: install reportlab")

    def _safe_pdf_text(value: Any) -> str:
        text = str(value or "")
        normalized = unicodedata.normalize("NFKD", text)
        return normalized.encode("latin-1", "ignore").decode("latin-1")

    buffer = BytesIO()
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))

    margin_left = 18
    margin_right = 18
    margin_top = 20
    margin_bottom = 20

    table_top = page_height - 120
    row_height = 16
    fixed_col_widths = [140, 115, 52, 72]  # parQ reduced to free horizontal space
    notes_width = 168
    preferred_date_col_width = 14

    def _draw_header_block(selected: ReportClass):
        y = page_height - margin_top
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left, y, _safe_pdf_text("Modalidade:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 58, y, _safe_pdf_text("Natação"))
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left + 250, y, _safe_pdf_text("PREFEITURA MUNICIPAL DE VINHEDO"))

        y -= 14
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left, y, _safe_pdf_text("Local:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 58, y, _safe_pdf_text("Piscina Bela Vista"))
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left + 250, y, _safe_pdf_text("SECRETARIA DE ESPORTE E LAZER"))

        y -= 14
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left, y, _safe_pdf_text("Professor:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 58, y, _safe_pdf_text(selected.professor or ""))

        y -= 14
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left, y, _safe_pdf_text("Turma:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 58, y, _safe_pdf_text(selected.turma or ""))
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left + 250, y, _safe_pdf_text("Nível:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 295, y, _safe_pdf_text(selected.nivel or ""))

        y -= 14
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left, y, _safe_pdf_text("Horário:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 58, y, _safe_pdf_text(_format_horario(selected.horario or "")))
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(margin_left + 250, y, _safe_pdf_text("Mês:"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin_left + 295, y, _safe_pdf_text(_format_month_label(month)))

    def _build_columns(day_chunk: List[str]) -> List[tuple[str, float]]:
        available_for_days = page_width - margin_left - margin_right - sum(fixed_col_widths) - notes_width
        if day_chunk:
            fit_width = available_for_days / len(day_chunk)
            dynamic_width = min(fit_width, preferred_date_col_width)
            notes_col_width = notes_width + max(0.0, available_for_days - (dynamic_width * len(day_chunk)))
        else:
            dynamic_width = available_for_days
            notes_col_width = notes_width
        columns: List[tuple[str, float]] = [
            ("Nome", fixed_col_widths[0]),
            ("Whatsapp", fixed_col_widths[1]),
            ("parQ", fixed_col_widths[2]),
            ("Aniversário", fixed_col_widths[3]),
        ]
        for day in day_chunk:
            columns.append((day, dynamic_width))
        columns.append(("Anotações", notes_col_width))
        return columns

    def _draw_grid_page(selected: ReportClass, rows: List[Dict[str, Any]], day_chunk: List[str], day_range_label: str):
        _draw_header_block(selected)
        if day_range_label:
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(margin_left, table_top + 6, day_range_label)

        columns = _build_columns(day_chunk)
        x_positions = [margin_left]
        for _, col_width in columns:
            x_positions.append(x_positions[-1] + col_width)

        y_top = table_top
        y_bottom = margin_bottom
        y = y_top

        pdf.setFont("Helvetica-Bold", 7)
        for col_idx, (label, _) in enumerate(columns):
            x0 = x_positions[col_idx]
            x1 = x_positions[col_idx + 1]
            pdf.rect(x0, y - row_height, x1 - x0, row_height)
            text = _safe_pdf_text(label or "")
            if col_idx in (0, len(columns) - 1):
                pdf.drawString(x0 + 2, y - 11, text[:28])
            else:
                # Personal data columns: whatsapp 18, parQ 10, aniversário 12
                if label == "Whatsapp":
                    max_chars = 18
                elif label == "parQ":
                    max_chars = 10
                else:
                    max_chars = 12
                pdf.drawCentredString((x0 + x1) / 2, y - 11, text[:max_chars])

        y -= row_height
        pdf.setFont("Helvetica", 7)

        for row in rows:
            if y - row_height < y_bottom:
                return False
            for col_idx, (label, _) in enumerate(columns):
                x0 = x_positions[col_idx]
                x1 = x_positions[col_idx + 1]
                pdf.rect(x0, y - row_height, x1 - x0, row_height)

                value = ""
                if label == "Nome":
                    value = _safe_pdf_text(row.get("nome") or "")
                elif label == "Whatsapp":
                    value = _safe_pdf_text(row.get("whatsapp") or "")
                elif label == "parQ":
                    value = _safe_pdf_text(row.get("parq") or "")
                elif label == "Aniversário":
                    value = _safe_pdf_text(row.get("data_nascimento") or "")
                elif label == "Anotações":
                    value = _safe_pdf_text(row.get("anotacoes") or "")
                elif label in day_chunk:
                    value = _safe_pdf_text((row.get("historico") or {}).get(label, ""))

                if col_idx in (0, len(columns) - 1):
                    pdf.drawString(x0 + 2, y - 11, value[:42])
                else:
                    # Personal data columns: whatsapp 18, parQ 10, aniversário 12
                    if label == "Whatsapp":
                        max_chars = 18
                    elif label == "parQ":
                        max_chars = 10
                    else:
                        max_chars = 12
                    pdf.drawCentredString((x0 + x1) / 2, y - 11, value[:max_chars])

            y -= row_height

        return True

    for report_idx, selected in enumerate(selected_reports):
        details_map = _build_import_student_details_map(
            session=session,
            turma=selected.turma,
            horario=selected.horario,
            professor=selected.professor,
        )

        class_days = _sort_report_days([
            day for aluno in selected.alunos for day in (aluno.historico or {}).keys()
        ])
        available_for_days = page_width - margin_left - margin_right - sum(fixed_col_widths) - notes_width
        max_date_slots_fit = max(1, int(available_for_days // preferred_date_col_width))
        if class_days:
            day_chunks = [
                class_days[idx: idx + max_date_slots_fit]
                for idx in range(0, len(class_days), max_date_slots_fit)
            ]
        else:
            day_chunks = [[]]

        students_sorted = sorted(selected.alunos, key=lambda item: item.nome)
        rows: List[Dict[str, Any]] = []
        for aluno in students_sorted:
            meta = details_map.get(_normalize_text(aluno.nome), {})

            rows.append(
                {
                    "nome": aluno.nome or "",
                    "whatsapp": meta.get("whatsapp") or "",
                    "parq": (meta.get("data_atestado") or "Com Atestado") if meta.get("atestado") else (meta.get("parq") or ""),
                    "data_nascimento": meta.get("data_nascimento") or "",
                    "historico": aluno.historico or {},
                    "anotacoes": aluno.anotacoes or "",
                }
            )

        if not rows:
            rows = [{"nome": "", "whatsapp": "", "parq": "", "data_nascimento": "", "historico": {}, "anotacoes": ""}]

        rows_per_page = int((table_top - margin_bottom) // row_height) - 1
        first_page_global = report_idx == 0
        for chunk_idx, day_chunk in enumerate(day_chunks):
            remaining = rows
            range_label = ""
            if day_chunk:
                range_label = f"Datas: {day_chunk[0]} a {day_chunk[-1]}"
            while remaining:
                if not first_page_global:
                    pdf.showPage()
                current_rows = remaining[:rows_per_page]
                remaining = remaining[rows_per_page:]
                label = range_label
                if len(day_chunks) > 1:
                    label = f"Bloco {chunk_idx + 1}/{len(day_chunks)} - {range_label}" if range_label else f"Bloco {chunk_idx + 1}/{len(day_chunks)}"
                _draw_grid_page(
                    selected=selected,
                    rows=current_rows,
                    day_chunk=day_chunk,
                    day_range_label=label,
                )
                first_page_global = False

    pdf.save()
    buffer.seek(0)
    return buffer.getvalue()

def _format_month_label(month: Optional[str]) -> str:
    if not month or "-" not in month:
        return str(month or "")
    year, month_number = month.split("-", 1)
    month_names = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    try:
        idx = max(1, min(12, int(month_number))) - 1
    except Exception:
        return str(month)
    return f"{month_names[idx]}/{year}"

def _copy_cell_style(ws, source_row: int, source_col: int, target_row: int, target_col: int) -> None:
    src = ws.cell(row=source_row, column=source_col)
    dst = ws.cell(row=target_row, column=target_col)
    dst._style = copy(src._style)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)
    dst.fill = copy(src.fill)
    dst.font = copy(src.font)
    dst.border = copy(src.border)

def _build_import_student_details_map(
    session: Session,
    turma: str,
    horario: str,
    professor: str,
) -> Dict[str, Dict[str, Any]]:
    classes = session.exec(select(models.ImportClass)).all()
    target = None
    turma_norm = _normalize_text(turma)
    horario_norm = _normalize_text(horario)
    professor_norm = _normalize_text(professor)
    for cls in classes:
        cls_label = cls.turma_label or cls.codigo or ""
        if (
            _normalize_text(cls_label) == turma_norm
            and _normalize_text(cls.horario or "") == horario_norm
            and _normalize_text(cls.professor or "") == professor_norm
        ):
            target = cls
            break

    details: Dict[str, Dict[str, Any]] = {}
    if not target:
        return details

    students = session.exec(
        select(models.ImportStudent).where(models.ImportStudent.class_id == target.id)
    ).all()
    for student in students:
        details[_normalize_text(student.nome)] = {
            "whatsapp": _format_whatsapp(student.whatsapp),
            "parq": student.parq or "",
            "atestado": bool(student.atestado),
            "data_nascimento": student.data_nascimento or "",
            "data_atestado": student.data_atestado or "",
        }
    return details

def _attendance_log_lookup_keys(item: Dict[str, Any]) -> List[str]:
    turma_codigo = str(item.get("turmaCodigo") or "").strip()
    turma_label = str(item.get("turmaLabel") or "").strip()
    horario = _normalize_horario_key(item.get("horario") or "")
    professor = str(item.get("professor") or "").strip()

    def _horario_variants(value: str) -> List[str]:
        raw = str(value or "").strip()
        digits = "".join(ch for ch in raw if ch.isdigit())
        variants: List[str] = []
        for candidate in [raw, digits, f"{digits[:2]}:{digits[2:4]}" if len(digits) >= 4 else ""]:
            token = str(candidate or "").strip()
            if token and token not in variants:
                variants.append(token)
        return variants or [""]

    def _professor_variants(value: str) -> List[str]:
        raw = str(value or "").strip()
        normalized = _normalize_text(raw)
        variants: List[str] = []
        for candidate in [raw, normalized]:
            token = str(candidate or "").strip()
            if token and token not in variants:
                variants.append(token)
        return variants or [""]

    def _turma_variants(value: str) -> List[str]:
        raw = str(value or "").strip()
        normalized = _normalize_text(raw)
        variants: List[str] = []
        for candidate in [raw, normalized]:
            token = str(candidate or "").strip()
            if token and token not in variants:
                variants.append(token)
        return variants or [""]

    keys: List[str] = []
    key_fields: List[List[str]] = [_horario_variants(horario), _professor_variants(professor)]

    for horario_key in key_fields[0]:
        for professor_key in key_fields[1]:
            if turma_codigo:
                keys.append("|".join(["codigo", turma_codigo, horario_key, professor_key]))
            for turma_key in _turma_variants(turma_label):
                keys.append("|".join(["label", turma_key, horario_key, professor_key]))

    return keys

def _load_latest_attendance_logs(month: Optional[str] = None) -> Dict[str, Dict[str, Any]]:
    # DB-first: lê do Supabase/PostgreSQL
    try:
        from app.database import engine as _db_engine
        from sqlmodel import Session as _DBSession
        with _DBSession(_db_engine) as _db:
            stmt = select(AttendanceLog)
            if month:
                stmt = stmt.where(AttendanceLog.mes == month)
            rows = _db.exec(stmt).all()
            if rows:
                latest: Dict[str, Dict[str, Any]] = {}
                for row in rows:
                    try:
                        registros = json.loads(row.registros_json or "[]")
                    except Exception:
                        registros = []
                    item = {
                        "turmaCodigo": row.turma_codigo,
                        "turmaLabel": row.turma_label,
                        "horario": row.horario,
                        "professor": row.professor,
                        "mes": row.mes,
                        "saved_at": row.saved_at,
                        "source": row.source,
                        "registros": registros,
                    }
                    keys = _attendance_log_lookup_keys(item)
                    saved_at_val = _saved_at_sort_key(row.saved_at)
                    for key in keys:
                        if key not in latest:
                            latest[key] = item
                            continue
                        existing_saved = _saved_at_sort_key(latest[key].get("saved_at"))
                        if saved_at_val > existing_saved:
                            latest[key] = item
                return latest
    except Exception:
        pass

    # Fallback: lê do JSON (dados históricos)
    file_path = os.path.join(DATA_DIR, "baseChamada.json")
    items = _load_json_list(file_path)
    latest: Dict[str, Dict[str, Any]] = {}
    for item in items:
        if month and str(item.get("mes") or "") != month:
            continue
        keys = _attendance_log_lookup_keys(item)
        if not keys:
            continue
        saved_at = _saved_at_sort_key(item.get("saved_at"))
        for key in keys:
            if key not in latest:
                latest[key] = item
                continue
            existing_saved = _saved_at_sort_key(latest[key].get("saved_at"))
            if saved_at > existing_saved:
                latest[key] = item
    return latest

@app.get("/exclusions")
def list_exclusions():
    with EXCLUSIONS_FILE_LOCK:
        return _read_exclusions_state(clean=True)


@app.get("/exclusions/backups")
def list_exclusions_backups(limit: int = Query(default=30, ge=1, le=200)):
    with EXCLUSIONS_FILE_LOCK:
        current_items = _read_exclusions_state(clean=True)
        backups = _list_exclusions_backups(limit=limit)
    return {
        "ok": True,
        "current_count": len(current_items),
        "backups": [{k: v for k, v in item.items() if k != "path"} for item in backups],
    }


@app.post("/exclusions/snapshot")
def snapshot_exclusions_state():
    with EXCLUSIONS_FILE_LOCK:
        current_items = _read_exclusions_state(clean=True)
        _save_json_list(_exclusions_file_path(), current_items)
        backups = _list_exclusions_backups(limit=1)
    return {
        "ok": True,
        "count": len(current_items),
        "backup_file": backups[0]["file"] if backups else None,
    }


@app.post("/exclusions/recover")
def recover_exclusions(payload: ExclusionsRecoverPayload):
    with EXCLUSIONS_FILE_LOCK:
        current_items = _read_exclusions_state(clean=True)
        backups = _list_exclusions_backups(limit=200)

        target: Optional[Dict[str, Any]] = None
        requested_backup = str(payload.backup_file or "").strip()
        if requested_backup:
            target = next((item for item in backups if item.get("file") == requested_backup), None)
            if target is None:
                raise HTTPException(status_code=404, detail="Backup de exclusoes nao encontrado")
        else:
            valid = [item for item in backups if int(item.get("count") or 0) >= max(0, int(payload.expected_min_items or 0))]
            if valid:
                target = max(valid, key=lambda item: (int(item.get("count") or 0), str(item.get("modified_at") or "")))

        if target is None:
            raise HTTPException(status_code=409, detail="Nenhum backup elegivel para recuperacao")

        target_items = _clean_exclusions_list(_load_json_list(str(target.get("path") or "")))
        if not target_items:
            raise HTTPException(status_code=409, detail="Backup selecionado nao possui exclusoes")

        if current_items and not payload.force:
            raise HTTPException(
                status_code=409,
                detail="Lista atual nao esta vazia. Envie force=true para sobrescrever ou merge=true para combinar",
            )

        if payload.merge:
            final_items = _merge_exclusions(current_items, target_items)
        else:
            final_items = target_items

        _write_exclusions_state(final_items, clean=True)

    return {
        "ok": True,
        "restored_from": target.get("file"),
        "before": len(current_items),
        "after": len(final_items),
        "backup_count": int(target.get("count") or 0),
    }

@app.post("/exclusions")
def add_exclusion(entry: ExclusionEntry):
    with EXCLUSIONS_FILE_LOCK:
        items = _read_exclusions_state(clean=True)
        payload = _normalize_exclusion_item(entry.dict())
        if not payload.get("dataExclusao"):
            payload["dataExclusao"] = pd.Timestamp.utcnow().strftime("%d/%m/%Y")

        updated = False
        for idx, item in enumerate(items):
            if _resolve_exclusion_match(item, entry):
                items[idx] = {**item, **payload}
                updated = True
                break
        if not updated:
            items.append(payload)
        _write_exclusions_state(items, clean=True)
        return {"ok": True, "updated": updated}


@app.post("/exclusions/bulk")
def bulk_upsert_exclusions(payload: ExclusionsBulkPayload):
    with EXCLUSIONS_FILE_LOCK:
        existing_items = [] if payload.replace else _read_exclusions_state(clean=True)

        updated = 0
        added = 0
        skipped = 0

        for entry in payload.items or []:
            normalized_entry = _normalize_exclusion_item(entry.dict())
            if not normalized_entry.get("dataExclusao"):
                normalized_entry["dataExclusao"] = pd.Timestamp.utcnow().strftime("%d/%m/%Y")

            uid = str(normalized_entry.get("student_uid") or normalized_entry.get("studentUid") or "").strip()
            item_id = str(normalized_entry.get("id") or "").strip()
            nome = _normalize_text(normalized_entry.get("nome") or normalized_entry.get("Nome") or "")
            if not uid and not item_id and not nome:
                skipped += 1
                continue

            match_index = next(
                (idx for idx, existing in enumerate(existing_items) if _exclusion_records_match(existing, normalized_entry)),
                -1,
            )
            if match_index >= 0:
                existing_items[match_index] = {**existing_items[match_index], **normalized_entry}
                updated += 1
            else:
                existing_items.append(normalized_entry)
                added += 1

        cleaned = _write_exclusions_state(existing_items, clean=True)

        return {
            "ok": True,
            "replace": payload.replace,
            "received": len(payload.items or []),
            "added": added,
            "updated": updated,
            "skipped": skipped,
            "total": len(cleaned),
        }

@app.post("/exclusions/restore")
def restore_exclusion(entry: ExclusionEntry):
    with EXCLUSIONS_FILE_LOCK:
        # NOTE: Do NOT clean/filter here - preserve all exclusion records
        items = _read_exclusions_state(clean=False)
        restored: Optional[Dict[str, Any]] = None
        remaining: List[Dict[str, Any]] = []
        for item in items:
            if restored is None and _resolve_exclusion_match(item, entry):
                restored = item
                continue
            remaining.append(item)
        _write_exclusions_state(remaining, clean=False)
        if restored is None:
            raise HTTPException(status_code=404, detail="Exclusion not found")
        return {"ok": True, "restored": restored}

@app.post("/exclusions/delete")
def delete_exclusion(entry: ExclusionEntry):
    with EXCLUSIONS_FILE_LOCK:
        # NOTE: Do NOT clean/filter here - preserve all exclusion records
        items = _read_exclusions_state(clean=False)
        remaining: List[Dict[str, Any]] = []
        deleted = False
        for item in items:
            if not deleted and _resolve_exclusion_match(item, entry):
                deleted = True
                continue
            remaining.append(item)
        _write_exclusions_state(remaining, clean=False)
        if not deleted:
            raise HTTPException(status_code=404, detail="Exclusion not found")
        return {"ok": True}

@app.get("/filters", response_model=ReportsFilterOut)
def get_report_filters(session: Session = Depends(get_session)) -> ReportsFilterOut:
    classes = session.exec(select(models.ImportClass)).all()
    turmas = sorted({(c.turma_label or c.codigo or "").strip() for c in classes if (c.turma_label or c.codigo)})
    horarios = sorted({(c.horario or "").strip() for c in classes if c.horario})
    professores = sorted({(c.professor or "").strip() for c in classes if c.professor})

    months = sorted({str(item.get("mes") or "").strip() for item in _load_json_list(os.path.join(DATA_DIR, "baseChamada.json")) if item.get("mes")})
    years = sorted({m.split("-")[0] for m in months if "-" in m})
    return ReportsFilterOut(turmas=turmas, horarios=horarios, professores=professores, meses=months, anos=years)

@app.get("/academic-calendar")
def get_academic_calendar(month: Optional[str] = None):
    state = _load_academic_calendar_state()
    events = state.get("events") or []
    bank_hours = state.get("bankHours") or []

    if month:
        try:
            month_start, month_end = _month_bounds(month)
            events = [
                item
                for item in events
                if month_start <= str(item.get("date") or "") <= month_end
            ]
            bank_hours = [
                item
                for item in bank_hours
                if month_start <= str(item.get("date") or "") <= month_end
            ]
        except Exception:
            pass

    return {
        "settings": state.get("settings"),
        "events": events,
        "bankHours": bank_hours,
    }

@app.put("/academic-calendar/settings")
def save_academic_calendar_settings(payload: AcademicCalendarSettingsPayload):
    state = _load_academic_calendar_state()
    state["settings"] = {
        "schoolYear": payload.schoolYear,
        "inicioAulas": payload.inicioAulas,
        "feriasInvernoInicio": payload.feriasInvernoInicio,
        "feriasInvernoFim": payload.feriasInvernoFim,
        "terminoAulas": payload.terminoAulas,
        "updated_at": datetime.utcnow().isoformat(),
    }
    _save_academic_calendar_state(state)
    return {"ok": True, "settings": state["settings"]}

@app.post("/academic-calendar/events")
def save_academic_calendar_event(payload: AcademicCalendarEventPayload):
    if payload.type not in {"feriado", "ponte", "reuniao", "evento"}:
        raise HTTPException(status_code=400, detail="Tipo de evento inválido")

    if payload.type == "reuniao" and not payload.allDay:
        if not payload.startTime or not payload.endTime:
            raise HTTPException(status_code=400, detail="Reunião por período exige horário de início e término")

    if payload.type == "evento":
        if not payload.startTime or not payload.endTime:
            raise HTTPException(status_code=400, detail="Evento exige horário de início e término")

    state = _load_academic_calendar_state()
    events = state.get("events") or []
    bank_hours = state.get("bankHours") or []

    event_id = str(uuid.uuid4())
    event = {
        "id": event_id,
        "date": payload.date,
        "type": payload.type,
        "allDay": bool(payload.allDay),
        "startTime": payload.startTime or "",
        "endTime": payload.endTime or "",
        "description": payload.description or "",
        "teacher": payload.teacher or "",
        "created_at": datetime.utcnow().isoformat(),
    }
    events.append(event)

    if payload.type == "evento":
        hours = _hours_from_interval(payload.startTime or "", payload.endTime or "")
        if hours > 0:
            bank_hours.append(
                {
                    "id": str(uuid.uuid4()),
                    "eventId": event_id,
                    "date": payload.date,
                    "teacher": payload.teacher or "",
                    "description": payload.description or "Evento",
                    "startTime": payload.startTime or "",
                    "endTime": payload.endTime or "",
                    "hours": hours,
                    "created_at": datetime.utcnow().isoformat(),
                }
            )

    state["events"] = events
    state["bankHours"] = bank_hours
    _save_academic_calendar_state(state)
    return {"ok": True, "event": event}

@app.delete("/academic-calendar/events/{event_id}")
def delete_academic_calendar_event(event_id: str):
    state = _load_academic_calendar_state()
    events = state.get("events") or []
    bank_hours = state.get("bankHours") or []

    existing_len = len(events)
    state["events"] = [item for item in events if str(item.get("id")) != str(event_id)]
    if len(state["events"]) == existing_len:
        raise HTTPException(status_code=404, detail="Evento não encontrado")

    state["bankHours"] = [item for item in bank_hours if str(item.get("eventId")) != str(event_id)]
    _save_academic_calendar_state(state)
    return {"ok": True}

@app.get("/reports", response_model=List[ReportClass])
def get_reports(month: Optional[str] = None, session: Session = Depends(get_session)) -> List[ReportClass]:
    classes = session.exec(select(models.ImportClass)).all()
    students = session.exec(select(models.ImportStudent)).all()
    excluded_items = _read_exclusions_state(clean=True)
    uid_registry = _load_student_uid_registry()
    uid_registry_changed = False

    students_by_class: Dict[int, List[models.ImportStudent]] = {}
    for student in students:
        students_by_class.setdefault(student.class_id, []).append(student)

    latest_logs = _load_latest_attendance_logs(month)
    report: List[ReportClass] = []

    for cls in classes:
        turma_key = (cls.codigo or cls.turma_label or "").strip()
        turma_label = (cls.turma_label or cls.codigo or "").strip()
        horario = (cls.horario or "").strip()
        professor = (cls.professor or "").strip()

        horario_digits = "".join(ch for ch in horario if ch.isdigit())
        horario_variants = [h for h in [horario, horario_digits, f"{horario_digits[:2]}:{horario_digits[2:4]}" if len(horario_digits) >= 4 else ""] if h]
        professor_variants = [p for p in [professor, _normalize_text(professor)] if p]
        if not horario_variants:
            horario_variants = [""]
        if not professor_variants:
            professor_variants = [""]

        turma_variants = [t for t in [turma_label, _normalize_text(turma_label)] if t]
        composite_keys: List[str] = []
        if turma_key:
            for h in horario_variants:
                for p in professor_variants:
                    composite_keys.append(f"codigo|{turma_key}|{h}|{p}")
        for turma_candidate in turma_variants:
            for h in horario_variants:
                for p in professor_variants:
                    composite_keys.append(f"label|{turma_candidate}|{h}|{p}")

        log_entry = None
        for key in composite_keys:
            log_entry = latest_logs.get(key)
            if log_entry:
                break

        class_roster = students_by_class.get(cls.id, [])
        name_to_student_meta: Dict[str, Dict[str, str]] = {}
        excluded_names = set()

        for student in class_roster:
            student_uid, changed = _ensure_student_uid_for_student(student, registry=uid_registry)
            uid_registry_changed = uid_registry_changed or changed
            normalized_name = _normalize_text(student.nome)
            if normalized_name and normalized_name not in name_to_student_meta:
                name_to_student_meta[normalized_name] = {
                    "id": str(student.id),
                    "student_uid": student_uid,
                }
            if _is_student_excluded_for_report(student, cls, excluded_items, student_uid=student_uid):
                excluded_names.add(normalized_name)

        class_students: List[ReportStudent] = []

        if log_entry:
            registros = log_entry.get("registros") or []
            for record in registros:
                nome = str(record.get("aluno_nome") or "").strip()
                normalized_name = _normalize_text(nome)
                if normalized_name in excluded_names:
                    continue
                attendance = record.get("attendance") or {}
                presencas = 0
                faltas = 0
                justificativas = 0
                historico: Dict[str, str] = {}
                for date_key, value in attendance.items():
                    mapped = _map_attendance_value(str(value))
                    if mapped == "c":
                        presencas += 1
                    elif mapped == "f":
                        faltas += 1
                    elif mapped == "j":
                        justificativas += 1
                    day_key = _report_day_key(date_key)
                    if day_key:
                        historico[day_key] = mapped

                justifications_raw = record.get("justifications") or {}
                justifications: Dict[str, str] = {}
                if isinstance(justifications_raw, dict):
                    for date_key, reason in justifications_raw.items():
                        normalized_date = _normalize_date_key(date_key)
                        normalized_reason = str(reason or "").strip()
                        if normalized_date and normalized_reason:
                            justifications[normalized_date] = normalized_reason

                notes_raw = record.get("notes") or []
                notes: List[str] = []
                if isinstance(notes_raw, list):
                    notes = [str(note or "").strip() for note in notes_raw if str(note or "").strip()]

                total = presencas + faltas + justificativas
                frequencia = round(((presencas + justificativas) / total) * 100, 1) if total else 0.0
                student_meta = name_to_student_meta.get(normalized_name, {})
                class_students.append(
                    ReportStudent(
                        id=student_meta.get("id") or nome or "0",
                        student_uid=student_meta.get("student_uid") or None,
                        nome=_to_proper_case(nome),
                        presencas=presencas,
                        faltas=faltas,
                        justificativas=justificativas,
                        frequencia=frequencia,
                        historico=historico,
                        justifications=justifications,
                        notes=notes,
                    )
                )
        else:
            for student in class_roster:
                normalized_name = _normalize_text(student.nome)
                if normalized_name in excluded_names:
                    continue
                student_meta = name_to_student_meta.get(normalized_name, {})
                class_students.append(
                    ReportStudent(
                        id=str(student.id),
                        student_uid=student_meta.get("student_uid") or None,
                        nome=_to_proper_case(student.nome),
                        presencas=0,
                        faltas=0,
                        justificativas=0,
                        frequencia=0.0,
                        historico={},
                        justifications={},
                        notes=[],
                    )
                )

        class_students.sort(key=lambda s: s.nome)
        report.append(
            ReportClass(
                turma=turma_label or turma_key,
                turmaCodigo=cls.codigo or "",
                horario=cls.horario or "",
                professor=cls.professor or "",
                nivel=cls.nivel or "",
                hasLog=bool(log_entry),
                alunos=class_students,
            )
        )

    report.sort(key=lambda c: (c.turma, c.horario))
    if uid_registry_changed:
        _save_student_uid_registry(uid_registry)
    return report

@app.post("/reports")
def generate_report(payload: Dict[str, Any], session: Session = Depends(get_session)):
    month = str(payload.get("month") or payload.get("mes") or "").strip() or None
    return get_reports(month=month, session=session)

@app.post("/reports/consolidated")
def generate_consolidated_report(payload: Dict[str, Any], session: Session = Depends(get_session)):
    return generate_report(payload, session=session)

@app.post("/reports/excel")
def generate_excel_report(payload: Dict[str, Any], session: Session = Depends(get_session)):
    return generate_report(payload, session=session)

@app.post("/reports/excel-file")
def generate_excel_report_file(payload: ExcelExportPayload, session: Session = Depends(get_session)):
    month = str(payload.month or "").strip() or None
    reports = get_reports(month=month, session=session)
    selected_reports = _resolve_export_targets(payload, reports)

    workbook = _build_excel_export_workbook(selected_reports=selected_reports, month=month, session=session)

    export_dir = os.path.join(DATA_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    safe_month = (month or "sem-mes").replace("/", "-")
    output_name = f"Relatorio_Multiturmas_{safe_month}.xlsx"
    output_path = os.path.join(export_dir, output_name)
    workbook.save(output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=output_name,
    )

@app.post("/reports/chamada-pdf-file")
def generate_chamada_pdf_file(payload: ExcelExportPayload, session: Session = Depends(get_session)):
    month = str(payload.month or "").strip() or None
    reports = get_reports(month=month, session=session)
    selected_reports = _resolve_export_targets(payload, reports)

    pdf_bytes = _build_chamada_pdf(selected_reports=selected_reports, month=month, session=session)

    export_dir = os.path.join(DATA_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    safe_month = (month or "sem-mes").replace("/", "-")
    output_name = f"Relatorio_Multiturmas_{safe_month}.pdf"
    output_path = os.path.join(export_dir, output_name)
    with open(output_path, "wb") as f:
        f.write(pdf_bytes)

    return FileResponse(
        output_path,
        media_type="application/pdf",
        filename=output_name,
    )


def _build_vacancies_workbook(payload: VacancyExportPayload):
    template_path = os.path.join(DATA_DIR, "templates", "vagasTemplate.xlsx")

    if os.path.exists(template_path):
        workbook = load_workbook(template_path)
        ws = workbook.active
    else:
        from openpyxl import Workbook

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Vagas"

    ws["A1"] = "Relatório de Vagas"
    ws["A2"] = f"Gerado em {payload.generatedAt or datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"] = (
        f"Totais: Lotação {payload.summary.totalLotacao}/{payload.summary.totalCapacidade}"
        f" | Vagas {payload.summary.totalVagas} | Excesso {payload.summary.totalExcesso}"
    )

    col_starts = [1, 6, 11]
    row_starts = [5, 11, 17, 23, 28]
    slots_per_sheet = len(col_starts) * len(row_starts)

    def _normalize_label(value: Any) -> str:
        return _normalize_text(str(value or ""))

    def _format_period_label(value: str) -> str:
        normalized = _normalize_text(value)
        if "ter" in normalized and "qui" in normalized:
            return "Ter/Qui"
        if "qua" in normalized and "sex" in normalized:
            return "Qua/Sex"
        return str(value or "").strip()

    def _slot_detail_capacity(row_idx: int, col_idx: int) -> int:
        # Layout fixo do template:
        # linhas 1-3: 2 detalhes em todas as colunas
        # linha 4: 1 detalhe em todas as colunas
        # linha 5: col 1 com 1 detalhe; col 2 e 3 com 2 detalhes
        if row_idx <= 2:
            return 2
        if row_idx == 3:
            return 1
        return 1 if col_idx == 0 else 2

    for slot in range(slots_per_sheet):
        row_idx = slot // len(col_starts)
        col_idx = slot % len(col_starts)
        row_start = row_starts[row_idx]
        col_start = col_starts[col_idx]
        detail_capacity = _slot_detail_capacity(row_idx, col_idx)
        detail_rows = [row_start + 1 + idx for idx in range(detail_capacity)]
        lotacao_row = row_start + 1 + detail_capacity
        vagas_row = lotacao_row + 1

        # Limpa somente células variáveis, preservando labels estáticos e estilos do template.
        ws.cell(row=row_start, column=col_start, value="")
        ws.cell(row=row_start, column=col_start + 1, value="")
        for row in detail_rows:
            ws.cell(row=row, column=col_start, value="")
            ws.cell(row=row, column=col_start + 1, value="")
            ws.cell(row=row, column=col_start + 2, value="")
        ws.cell(row=lotacao_row, column=col_start + 1, value="")
        ws.cell(row=vagas_row, column=col_start + 1, value="")
        ws.cell(row=vagas_row, column=col_start + 3, value="")

        if slot >= len(payload.blocks):
            continue

        block = payload.blocks[slot]
        ws.cell(row=row_start, column=col_start, value=_format_horario(block.horario or ""))
        ws.cell(row=row_start, column=col_start + 1, value=_format_period_label(block.periodoLabel or ""))

        visible_rows = block.rows[: len(detail_rows)]
        extra_rows = block.rows[len(detail_rows) :]
        for idx, row in enumerate(detail_rows):
            detail = visible_rows[idx] if idx < len(visible_rows) else None
            ws.cell(row=row, column=col_start, value=(f"{detail.nivel}:" if detail else ""))
            ws.cell(
                row=row,
                column=col_start + 1,
                value=(f"{detail.lotacao}/{detail.capacidade}" if detail else ""),
            )
            ws.cell(row=row, column=col_start + 2, value=(detail.professor if detail else ""))

        if extra_rows and detail_rows:
            last_detail_row = detail_rows[-1]
            extra_text = " | ".join(
                f"{item.nivel} {item.lotacao}/{item.capacidade} {item.professor}".strip()
                for item in extra_rows
            )
            current_prof = str(ws.cell(row=last_detail_row, column=col_start + 2).value or "").strip()
            ws.cell(
                row=last_detail_row,
                column=col_start + 2,
                value=(f"{current_prof} | {extra_text}" if current_prof else extra_text),
            )

        ws.cell(
            row=lotacao_row,
            column=col_start + 1,
            value=f"{block.lotacaoHorario}/{block.capacidadeHorario}",
        )
        ws.cell(row=lotacao_row, column=col_start + 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=vagas_row, column=col_start + 1, value=block.vagasDisponiveis)
        ws.cell(row=vagas_row, column=col_start + 3, value=block.excesso)
        ws.cell(row=vagas_row, column=col_start + 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=vagas_row, column=col_start + 3).alignment = Alignment(horizontal="left", vertical="center")

    return workbook


def _build_vacancies_pdf(payload: VacancyExportPayload) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF export unavailable: install reportlab")

    buffer = BytesIO()
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(24, page_height - 28, "Relatório de Vagas - Template")
    pdf.setFont("Helvetica", 8)
    generated_at = payload.generatedAt or datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pdf.drawString(24, page_height - 40, f"Gerado em {generated_at}")
    pdf.drawString(
        24,
        page_height - 50,
        (
            f"Totais: Lotação {payload.summary.totalLotacao}/{payload.summary.totalCapacidade}"
            f" | Vagas {payload.summary.totalVagas} | Excesso {payload.summary.totalExcesso}"
        ),
    )

    start_x = 24
    start_y = page_height - 70
    col_starts = [24, 190, 356]
    # Mirror the workbook's actual row heights: default rows are 12.75 pt and footer/spacer rows are 13.5 pt.
    row_starts = [start_y, start_y - 78.0, start_y - 156.0, start_y - 234.0, start_y - 299.25]
    block_width = 156
    body_row_height = 12.75
    footer_row_height = 13.5
    header_height = body_row_height
    left_padding = 4
    header_time_x = 4
    header_period_x = 34
    detail_ratio_x = 50
    detail_prof_x = 94

    slots_per_page = len(col_starts) * len(row_starts)
    blocks = payload.blocks[: slots_per_page]

    for slot in range(slots_per_page):
        row_idx = slot // len(col_starts)
        col_idx = slot % len(col_starts)
        x1 = col_starts[col_idx]
        y1 = row_starts[row_idx]

        if row_idx <= 2:
            detail_capacity = 2
        elif row_idx == 3:
            detail_capacity = 1
        else:
            detail_capacity = 1 if col_idx == 0 else 2
        block_height = body_row_height * (detail_capacity + 2) + footer_row_height
        x2 = x1 + block_width
        y2 = y1 - block_height

        pdf.setLineWidth(1.1)
        pdf.rect(x1, y2, block_width, block_height)

        if slot >= len(blocks):
            continue

        block = blocks[slot]

        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(x1 + header_time_x, y1 - 8, _format_horario(block.horario or ""))
        pdf.drawString(x1 + header_period_x, y1 - 8, block.periodoLabel or "")

        visible_rows = block.rows[:detail_capacity]
        extra_rows = block.rows[detail_capacity:]
        row_top = y1 - header_height
        for idx in range(detail_capacity):
            next_row = row_top - body_row_height
            detail = visible_rows[idx] if idx < len(visible_rows) else None
            detail_label = f"{detail.nivel}:" if detail else ""
            detail_ratio = f"{detail.lotacao}/{detail.capacidade}" if detail else ""
            detail_prof = str(detail.professor or "") if detail else ""

            if idx == detail_capacity - 1 and extra_rows:
                suffix = " | ".join(
                    f"{item.nivel} {item.lotacao}/{item.capacidade} {item.professor}".strip()
                    for item in extra_rows
                )
                detail_prof = f"{detail_prof} | {suffix}" if detail_prof else suffix

            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(x1 + left_padding, row_top - 8, detail_label)
            pdf.setFont("Helvetica", 8)
            pdf.drawString(x1 + detail_ratio_x, row_top - 8, detail_ratio)
            pdf.drawString(x1 + detail_prof_x, row_top - 8, detail_prof[:36])
            row_top = next_row

        lotacao_row = row_top - body_row_height
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(x1 + left_padding, row_top - 8, "Lotação:")
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x1 + detail_ratio_x, row_top - 8, f"{block.lotacaoHorario}/{block.capacidadeHorario}")

        footer_row = lotacao_row - footer_row_height
        middle_x = x1 + (block_width / 2)
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(x1 + left_padding, lotacao_row - 8, "Vagas:")
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x1 + 35, lotacao_row - 8, str(block.vagasDisponiveis))
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(middle_x + 4, lotacao_row - 8, "Excesso:")
        pdf.setFont("Helvetica", 8)
        pdf.drawString(middle_x + 42, lotacao_row - 8, str(block.excesso))

    pdf.save()
    buffer.seek(0)
    return buffer.getvalue()


@app.post("/reports/vacancies-excel-file")
def generate_vacancies_excel_file(payload: VacancyExportPayload):
    if not payload.blocks:
        raise HTTPException(status_code=400, detail="No vacancy data informed")

    workbook = _build_vacancies_workbook(payload)

    export_dir = os.path.join(DATA_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    output_name = f"Relatorio_Vagas_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    output_path = os.path.join(export_dir, output_name)
    workbook.save(output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=output_name,
    )


@app.post("/reports/vacancies-pdf-file")
def generate_vacancies_pdf_file(payload: VacancyExportPayload):
    if not payload.blocks:
        raise HTTPException(status_code=400, detail="No vacancy data informed")

    pdf_bytes = _build_vacancies_pdf(payload)

    export_dir = os.path.join(DATA_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    output_name = f"Relatorio_Vagas_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    output_path = os.path.join(export_dir, output_name)
    with open(output_path, "wb") as f:
        f.write(pdf_bytes)

    return FileResponse(
        output_path,
        media_type="application/pdf",
        filename=output_name,
    )


# --- Statistics aggregation endpoint ---
@app.get("/reports/statistics", response_model=List[StudentStatisticsOut])
def get_reports_statistics(session: Session = Depends(get_session)):
    """Aggregate per-student retention and per-level permanence from attendance logs.

    - firstPresence: primeira data com registro de presença ('Presente' ou 'Justificado').
    - retention: dias entre firstPresence e data de exclusão (se houver) ou hoje.
    - levels: para cada nível onde o aluno teve registros, retorna período e frequência.

    Nota: identificação de aluno é por nome normalizado (mesmo critério usado em relatórios atuais).
    """
    # load classes to map turmaCodigo/turmaLabel -> nivel
    overrides_applied = _apply_transfer_overrides(session)
    if overrides_applied > 0:
        session.commit()

    classes = session.exec(select(models.ImportClass)).all()
    class_by_code = {str(c.codigo or ""): c for c in classes}
    class_by_label_norm = {_normalize_text_fold(c.turma_label or ""): c for c in classes if str(c.turma_label or "").strip()}
    class_by_triple: Dict[Tuple[str, str, str], List[models.ImportClass]] = {}
    for c in classes:
        triple_key = (
            _normalize_text_fold(c.turma_label or ""),
            _normalize_horario_key(c.horario or ""),
            _normalize_text_fold(c.professor or ""),
        )
        if any(triple_key):
            class_by_triple.setdefault(triple_key, []).append(c)

    # map current active class level by student name from import tables (source of truth for current allocation)
    active_level_by_name: Dict[str, Dict[str, Any]] = {}
    import_students = session.exec(select(models.ImportStudent)).all()
    class_by_id = {int(c.id): c for c in classes if getattr(c, "id", None) is not None}
    for st in import_students:
        nome_raw = str(getattr(st, "nome", "") or "").strip()
        if not nome_raw:
            continue
        name_key = _normalize_text_fold(nome_raw)
        cls = class_by_id.get(int(st.class_id)) if getattr(st, "class_id", None) else None
        nivel_raw = str(getattr(cls, "nivel", "") or "").strip() if cls else ""
        candidate = {
            "nivel": nivel_raw,
            "student_id": int(getattr(st, "id", 0) or 0),
            "class_id": int(getattr(st, "class_id", 0) or 0),
            "dias_semana": str(getattr(cls, "dias_semana", "") or "") if cls else "",
            "turma_label": str(getattr(cls, "turma_label", "") or "") if cls else "",
            "codigo": str(getattr(cls, "codigo", "") or "") if cls else "",
            "data_nascimento": str(getattr(st, "data_nascimento", "") or ""),
            "nome": _to_proper_case(nome_raw),
        }
        existing = active_level_by_name.get(name_key)
        if not existing:
            active_level_by_name[name_key] = candidate
            continue
        # prefer entries with level, and then the latest import student id
        existing_has_level = bool(str(existing.get("nivel") or "").strip())
        candidate_has_level = bool(nivel_raw)
        if (candidate_has_level and not existing_has_level) or (
            candidate_has_level == existing_has_level and candidate["student_id"] >= int(existing.get("student_id") or 0)
        ):
            active_level_by_name[name_key] = candidate

    # load all attendance log entries (sorted to let the newest snapshot win)
    items = _load_json_list(os.path.join(DATA_DIR, "baseChamada.json"))
    items = sorted(items, key=lambda item: _saved_at_sort_key((item or {}).get("saved_at")))

    today = datetime.utcnow().date()
    allowed_days_map = _load_allowed_schedule_days(today)

    students: Dict[str, Dict[str, Any]] = {}

    def _ensure_student(name: str):
        key = _normalize_text_fold(name)
        if key not in students:
            students[key] = {
                "nome": _to_proper_case(name),
                "ids": set(),
                "attendance_by_date": {},  # date -> { status, nivel, turmaCodigo }
                "per_level": {},
                "event_by_key": {},  # deduped event map
                "first_presence": None,
                "last_presence": None,
                "exclusion_date": None,
            }
        return students[key]

    # iterate logs (order not guaranteed) — process dates in each registro
    for item in items:
        turma_codigo = str(item.get("turmaCodigo") or "").strip()
        turma_label = str(item.get("turmaLabel") or "").strip()
        turma_horario = _normalize_horario_key(item.get("horario") or "")
        turma_professor = _normalize_text(item.get("professor") or "")
        schedule_group = _infer_schedule_group(turma_label, turma_codigo)
        # resolve nivel from import classes
        nivel = ""
        cls = None
        ambiguous_level = False
        class_triple_key = (
            _normalize_text_fold(turma_label),
            turma_horario,
            turma_professor,
        )

        # Prefer class code when available (historically more stable than generic labels).
        if turma_codigo and turma_codigo in class_by_code:
            cls = class_by_code.get(turma_codigo)
        elif class_triple_key in class_by_triple:
            triple_candidates = class_by_triple.get(class_triple_key) or []
            if len(triple_candidates) == 1:
                cls = triple_candidates[0]
            elif len(triple_candidates) > 1:
                ambiguous_level = True
        elif turma_label:
            cls = class_by_label_norm.get(_normalize_text_fold(turma_label))

        # Generic day labels without class code can map to multiple levels over time.
        if not turma_codigo and schedule_group in {"tq", "qs"}:
            ambiguous_level = True

        if cls:
            nivel = str(cls.nivel or "")

        registros = item.get("registros") or []
        for record in registros:
            nome = str(record.get("aluno_nome") or "").strip()
            if not nome:
                continue
            st = _ensure_student(nome)
            attendance_map = record.get("attendance") or {}
            # iterate dates in chronological order
            for date_key in sorted(attendance_map.keys()):
                raw = str(attendance_map.get(date_key) or "").strip()
                mapped = _map_attendance_value(raw)
                # keep records for frequency calculation even if empty (we'll skip empty)
                if not date_key:
                    continue
                # ignore empty markers
                if mapped == "":
                    continue
                try:
                    parsed_d = datetime.strptime(date_key, "%Y-%m-%d").date()
                except Exception:
                    continue

                # dedupe snapshots: same student/day/schedule should count once (latest wins)
                if schedule_group in {"tq", "qs"}:
                    event_key = (date_key, schedule_group, turma_horario, turma_professor)
                else:
                    event_key = (
                        date_key,
                        _normalize_text(turma_codigo or turma_label),
                        turma_horario,
                        turma_professor,
                    )

                st["event_by_key"][event_key] = {
                    "date": parsed_d,
                    "status": mapped,
                    "nivel": nivel or "(sem-nivel)",
                    "ambiguous_level": bool(ambiguous_level),
                }

    # build per-student aggregates from deduped events
    for student_key, st in students.items():
        st["attendance_by_date"] = {}
        st["per_level"] = {}
        st["first_presence"] = None
        st["last_presence"] = None

        active_level_for_student = str((active_level_by_name.get(student_key) or {}).get("nivel") or "").strip()

        deduped_events = sorted(
            st.get("event_by_key", {}).values(),
            key=lambda event: event.get("date") or date.min,
        )

        for event in deduped_events:
            parsed_d = event.get("date")
            if not parsed_d:
                continue
            date_key = parsed_d.isoformat()
            mapped = str(event.get("status") or "")
            level_key = str(event.get("nivel") or "(sem-nivel)")
            if (level_key == "(sem-nivel)" or bool(event.get("ambiguous_level"))) and active_level_for_student:
                level_key = active_level_for_student

            st["attendance_by_date"][date_key] = {
                "status": mapped,
                "nivel": level_key,
                "turmaCodigo": "",
            }

            if mapped in {"c", "j"}:
                if st["first_presence"] is None or parsed_d < st["first_presence"]:
                    st["first_presence"] = parsed_d
                if st["last_presence"] is None or parsed_d > st["last_presence"]:
                    st["last_presence"] = parsed_d

            lvl = st["per_level"].setdefault(level_key, {
                "first": None,
                "last": None,
                "presencas": 0,
                "faltas": 0,
                "justificativas": 0,
            })

            if lvl["first"] is None or parsed_d < lvl["first"]:
                lvl["first"] = parsed_d
            if lvl["last"] is None or parsed_d > lvl["last"]:
                lvl["last"] = parsed_d

            if mapped == "c":
                lvl["presencas"] += 1
            elif mapped == "f":
                lvl["faltas"] += 1
            elif mapped == "j":
                lvl["justificativas"] += 1

    # load exclusions
    excluded = _read_exclusions_state(clean=True)
    for ex in excluded:
        nome = str(ex.get("nome") or "").strip()
        if not nome:
            continue
        key = _normalize_text_fold(nome)
        if key not in students:
            students[key] = {"nome": _to_proper_case(nome), "ids": set(), "attendance_by_date": {}, "per_level": {}, "event_by_key": {}, "first_presence": None, "last_presence": None, "exclusion_date": None}
        date_str = str(ex.get("dataExclusao") or "").strip()
        if date_str:
            try:
                parsed = datetime.strptime(date_str, "%d/%m/%Y").date()
                students[key]["exclusion_date"] = parsed
            except Exception:
                # keep as None on parse failure
                pass

    # build output list
    out: List[Dict[str, Any]] = []
    for key, st in sorted(students.items(), key=lambda t: t[1]["nome"].lower()):
        first = st.get("first_presence")
        last = st.get("last_presence")
        exclusion = st.get("exclusion_date")
        end_date = exclusion or today
        retention_days = 0
        if first and end_date:
            retention_days = max(0, (end_date - first).days)
        # determine current nivel: prefer active student allocation in import tables, fallback to latest attendance history
        current_nivel = None
        active_entry = active_level_by_name.get(key)
        if active_entry:
            resolved_active_level = str(active_entry.get("nivel") or "").strip()
            if resolved_active_level and resolved_active_level != "(sem-nivel)":
                current_nivel = resolved_active_level

        if not current_nivel and st.get("per_level"):
            candidates = [(k, v["last"]) for k, v in st["per_level"].items() if v.get("last")]
            if candidates:
                candidates.sort(key=lambda x: x[1], reverse=True)
                current_nivel = candidates[0][0]
                if current_nivel == "(sem-nivel)":
                    current_nivel = None

        def _find_level_values(level_name: str) -> Optional[Dict[str, Any]]:
            if not level_name:
                return None
            exact = st.get("per_level", {}).get(level_name)
            if exact:
                return exact
            normalized_target = _normalize_text(level_name)
            for stored_level_name, stored_values in st.get("per_level", {}).items():
                if _normalize_text(stored_level_name) == normalized_target:
                    return stored_values
            return None

        def _schedule_group_from_active_entry(entry: Dict[str, Any]) -> str:
            dias_semana = _normalize_text_fold(str(entry.get("dias_semana") or ""))
            if ("terca" in dias_semana and "quinta" in dias_semana) or "tq" in dias_semana:
                return "tq"
            if ("quarta" in dias_semana and "sexta" in dias_semana) or "qs" in dias_semana:
                return "qs"
            return _infer_schedule_group(str(entry.get("turma_label") or ""), str(entry.get("codigo") or ""))

        def _count_planned_days(entry: Dict[str, Any], start_date: date, end_date: date) -> int:
            if not start_date or not end_date or start_date > end_date:
                return 0
            group = _schedule_group_from_active_entry(entry)
            if group not in {"tq", "qs"}:
                return 0
            allowed_days = allowed_days_map.get(group, set())
            return sum(1 for day_key in allowed_days if start_date.isoformat() <= day_key <= end_date.isoformat())

        # Merge stray Adult buckets into current non-adult level before rendering levels list.
        if current_nivel and "adult" not in _normalize_text_fold(current_nivel):
            adult_bucket_keys = [
                lvl_name
                for lvl_name in list(st.get("per_level", {}).keys())
                if "adult" in _normalize_text_fold(lvl_name)
            ]
            if adult_bucket_keys:
                target_bucket = st.setdefault("per_level", {}).setdefault(current_nivel, {
                    "first": None,
                    "last": None,
                    "presencas": 0,
                    "faltas": 0,
                    "justificativas": 0,
                })
                for adult_key in adult_bucket_keys:
                    adult_bucket = st.get("per_level", {}).get(adult_key) or {}
                    adult_first = adult_bucket.get("first")
                    adult_last = adult_bucket.get("last")
                    if adult_first and (target_bucket.get("first") is None or adult_first < target_bucket.get("first")):
                        target_bucket["first"] = adult_first
                    if adult_last and (target_bucket.get("last") is None or adult_last > target_bucket.get("last")):
                        target_bucket["last"] = adult_last
                    target_bucket["presencas"] = int(target_bucket.get("presencas") or 0) + int(adult_bucket.get("presencas") or 0)
                    target_bucket["faltas"] = int(target_bucket.get("faltas") or 0) + int(adult_bucket.get("faltas") or 0)
                    target_bucket["justificativas"] = int(target_bucket.get("justificativas") or 0) + int(adult_bucket.get("justificativas") or 0)
                    st.get("per_level", {}).pop(adult_key, None)

        # build levels array
        levels_out: List[LevelHistoryOut] = []
        for lvl_name, vals in st.get("per_level", {}).items():
            first_d = vals.get("first")
            last_d = vals.get("last")
            pres = int(vals.get("presencas") or 0)
            falt = int(vals.get("faltas") or 0)
            just = int(vals.get("justificativas") or 0)
            total = pres + falt + just
            freq = round(((pres + just) / total) * 100, 1) if total else 0.0
            days = 0
            if first_d and last_d:
                days = max(0, (last_d - first_d).days + 1)
            name_out = None if lvl_name == "(sem-nivel)" else lvl_name
            levels_out.append(LevelHistoryOut(
                nivel=name_out or "",
                firstDate=first_d.isoformat() if first_d else None,
                lastDate=last_d.isoformat() if last_d else None,
                days=days,
                presencas=pres,
                faltas=falt,
                justificativas=just,
                frequencia=freq,
            ))

        if current_nivel:
            active_entry = active_level_by_name.get(key) or {}
            current_key = _normalize_text(current_nivel)
            current_first = None
            current_last = None
            current_pres = 0
            current_falt = 0
            current_just = 0
            current_total_days = 0

            def _aggregate_all_levels() -> Dict[str, Any]:
                agg_first = None
                agg_last = None
                agg_pres = 0
                agg_falt = 0
                agg_just = 0
                for lvl_name, vals in st.get("per_level", {}).items():
                    if lvl_name == "(sem-nivel)":
                        continue
                    lvl_first = vals.get("first")
                    lvl_last = vals.get("last")
                    if lvl_first and (agg_first is None or lvl_first < agg_first):
                        agg_first = lvl_first
                    if lvl_last and (agg_last is None or lvl_last > agg_last):
                        agg_last = lvl_last
                    agg_pres += int(vals.get("presencas") or 0)
                    agg_falt += int(vals.get("faltas") or 0)
                    agg_just += int(vals.get("justificativas") or 0)
                return {
                    "first": agg_first,
                    "last": agg_last,
                    "presencas": agg_pres,
                    "faltas": agg_falt,
                    "justificativas": agg_just,
                }

            current_level_vals = _find_level_values(current_nivel)
            current_first = current_level_vals.get("first") if current_level_vals else None
            current_last = current_level_vals.get("last") if current_level_vals else None

            if current_first is None:
                previous_last_dates = [vals.get("last") for lvl, vals in st.get("per_level", {}).items() if _normalize_text(lvl) != current_key and vals.get("last")]
                current_first = (max(previous_last_dates) + timedelta(days=1)) if previous_last_dates else first

            current_last = current_last or end_date

            # Count all recorded events in the current period, regardless of the bucket that produced them.
            period_entries = [
                entry
                for date_key, entry in st.get("attendance_by_date", {}).items()
                if date_key and current_first and current_last and current_first.isoformat() <= date_key <= current_last.isoformat()
            ]
            current_pres = sum(1 for entry in period_entries if str(entry.get("status") or "") == "c")
            current_falt = sum(1 for entry in period_entries if str(entry.get("status") or "") == "f")
            current_just = sum(1 for entry in period_entries if str(entry.get("status") or "") == "j")
            current_total_days = current_pres + current_falt + current_just

            if current_total_days == 0 and current_level_vals:
                current_pres = int(current_level_vals.get("presencas") or 0)
                current_falt = int(current_level_vals.get("faltas") or 0)
                current_just = int(current_level_vals.get("justificativas") or 0)
                current_total_days = current_pres + current_falt + current_just

            if current_total_days == 0:
                aggregate_vals = _aggregate_all_levels()
                current_first = aggregate_vals.get("first") or current_first or first
                current_last = aggregate_vals.get("last") or current_last or end_date
                current_pres = int(aggregate_vals.get("presencas") or 0)
                current_falt = int(aggregate_vals.get("faltas") or 0)
                current_just = int(aggregate_vals.get("justificativas") or 0)
                current_total_days = current_pres + current_falt + current_just

            if current_total_days < 0:
                current_total_days = 0

            current_frequency = round(((current_pres + current_just) / current_total_days) * 100, 1) if current_total_days else 0.0
            current_entry = LevelHistoryOut(
                nivel=current_nivel,
                firstDate=current_first.isoformat() if current_first else None,
                lastDate=current_last.isoformat() if current_last else None,
                days=current_total_days,
                presencas=current_pres,
                faltas=current_falt,
                justificativas=current_just,
                frequencia=current_frequency,
            )

            replaced = False
            for idx, entry in enumerate(levels_out):
                if _normalize_text(entry.nivel or "") == current_key:
                    levels_out[idx] = current_entry
                    replaced = True
                    break
            if not replaced:
                levels_out.append(current_entry)

        levels_out.sort(key=lambda lvl: (lvl.lastDate or "", lvl.firstDate or ""), reverse=True)
        out.append(StudentStatisticsOut(
            id=None,
            nome=_to_proper_case(st.get("nome") or key),
            firstPresence=first.isoformat() if first else None,
            lastPresence=last.isoformat() if last else None,
            exclusionDate=exclusion.isoformat() if exclusion else None,
            retentionDays=retention_days,
            currentNivel=current_nivel,
            levels=levels_out,
        ))
    return out

@app.post("/reports/parq-pdf-file")
def generate_parq_pdf_file_compat(payload: ExcelExportPayload, session: Session = Depends(get_session)):
    return generate_chamada_pdf_file(payload=payload, session=session)

def parse_bool(value: str) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "sim", "yes", "y"}


def _normalize_horario_value(raw: str) -> str:
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if not digits:
        return ""
    if len(digits) == 3:
        digits = f"0{digits}"
    if len(digits) > 4:
        digits = digits[:4]
    return digits.zfill(4)

def get_or_create_import_unit(session: Session, name: str) -> models.ImportUnit:
    stmt = select(models.ImportUnit).where(models.ImportUnit.name == name)
    unit = session.exec(stmt).first()
    if unit:
        return unit
    unit = models.ImportUnit(name=name)
    session.add(unit)
    session.flush()
    return unit

def get_or_create_import_class(session: Session, unit_id: int, codigo: str, horario: str) -> models.ImportClass | None:
    stmt = select(models.ImportClass).where(
        models.ImportClass.unit_id == unit_id,
        models.ImportClass.codigo == codigo,
        models.ImportClass.horario == horario,
    )
    return session.exec(stmt).first()

def get_or_create_import_student(session: Session, class_id: Optional[int], nome: str) -> models.ImportStudent | None:
    if class_id is None:
        stmt = select(models.ImportStudent).where(
            models.ImportStudent.class_id.is_(None),
            models.ImportStudent.nome == nome,
        )
    else:
        stmt = select(models.ImportStudent).where(
            models.ImportStudent.class_id == class_id,
            models.ImportStudent.nome == nome,
        )
    return session.exec(stmt).first()

def _find_import_class_by_triple(
    session: Session,
    turma: str,
    horario: str,
    professor: str,
) -> Optional[models.ImportClass]:
    turma_norm = _normalize_text(turma)
    horario_key = _normalize_horario_value(horario)
    professor_norm = _normalize_text(professor)
    if not turma_norm or not horario_key or not professor_norm:
        return None

    classes = session.exec(select(models.ImportClass)).all()
    for cls in classes:
        cls_horario = _normalize_horario_value(cls.horario or "")
        cls_professor = _normalize_text(cls.professor or "")
        cls_codigo = _normalize_text(cls.codigo or "")
        cls_label = _normalize_text(cls.turma_label or cls.codigo or "")
        turma_matches = turma_norm in {cls_codigo, cls_label}
        if turma_matches and cls_horario == horario_key and cls_professor == professor_norm:
            return cls
    return None

def _import_student_out(
    student: models.ImportStudent,
    uid_registry: Optional[Dict[str, str]] = None,
    preferred_uid: Optional[str] = None,
) -> ImportStudentOut:
    student_uid, _ = _ensure_student_uid_for_student(
        student,
        registry=uid_registry,
        preferred_uid=preferred_uid,
    )

    return ImportStudentOut(
        id=student.id or 0,
        student_uid=student_uid,
        class_id=student.class_id,
        nome=student.nome,
        whatsapp=student.whatsapp or "",
        data_nascimento=student.data_nascimento or "",
        data_atestado=student.data_atestado or "",
        categoria=student.categoria or "",
        genero=student.genero or "",
        parq=student.parq or "",
        atestado=bool(student.atestado),
    )

@app.post("/api/import-data", response_model=ImportResult)
async def import_data(
    file: UploadFile = File(...),
    apply_overrides: bool = Form(True),
    session: Session = Depends(get_session),
) -> ImportResult:
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="File must be .csv")

    content = await file.read()
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")

    try:
        dialect = csv.Sniffer().sniff(text, delimiters=[",", ";", "\t"])
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    if reader.fieldnames and len(reader.fieldnames) == 1:
        header_line = reader.fieldnames[0]
        if ";" in header_line:
            delimiter = ";"
        elif "\t" in header_line:
            delimiter = "\t"
        if delimiter != ",":
            reader = csv.DictReader(StringIO(text), delimiter=delimiter)

    # Personal columns always required; class columns are optional (enables personal-only CSV)
    personal_required = {
        "aluno_nome",
        "whatsapp",
        "data_nascimento",
        "data_atest",
        "categoria",
        "genero",
        "parq",
        "atestado",
    }
    class_columns = {"unidade", "turma_codigo", "horario", "professor", "nivel", "capacidade", "dias_semana", "aluno_turma"}

    if reader.fieldnames is None:
        raise HTTPException(status_code=400, detail="CSV header not found")

    fieldnames_stripped = {name.strip() for name in reader.fieldnames}
    missing_personal = personal_required.difference(fieldnames_stripped)
    if missing_personal:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(sorted(missing_personal))}")

    full_mode = class_columns.issubset(fieldnames_stripped)

    counters: Dict[str, int] = {
        "units_created": 0,
        "units_updated": 0,
        "classes_created": 0,
        "classes_updated": 0,
        "students_created": 0,
        "students_updated": 0,
    }
    rows_processed = 0

    try:
        for row in reader:
            aluno_nome = (row.get("aluno_nome") or "").strip()
            if not aluno_nome:
                continue

            rows_processed += 1

            if full_mode:
                unidade = (row.get("unidade") or "").strip()
                codigo = (row.get("turma_codigo") or "").strip()
                raw_horario = (row.get("horario") or "").strip()
                horario = _normalize_horario_value(raw_horario)

                if not unidade or not codigo or not horario:
                    # Missing class identifiers even in full mode — treat as personal
                    student = get_or_create_import_student(session, None, aluno_nome)
                    if student:
                        counters["students_updated"] += 1
                    else:
                        student = models.ImportStudent(class_id=None, nome=aluno_nome)
                        session.add(student)
                        counters["students_created"] += 1
                else:
                    unit = get_or_create_import_unit(session, unidade)
                    if unit.id:
                        counters["units_updated"] += 1
                    else:
                        counters["units_created"] += 1

                    class_obj = get_or_create_import_class(session, unit.id, codigo, horario)
                    if class_obj:
                        counters["classes_updated"] += 1
                    else:
                        class_obj = models.ImportClass(unit_id=unit.id, codigo=codigo, horario=horario)
                        session.add(class_obj)
                        session.flush()
                        counters["classes_created"] += 1

                    class_obj.horario = horario
                    class_obj.professor = (row.get("professor") or "").strip()
                    class_obj.nivel = (row.get("nivel") or "").strip()
                    class_obj.faixa_etaria = (row.get("faixa_etaria") or "").strip()
                    class_obj.capacidade = int((row.get("capacidade") or "0") or 0)
                    class_obj.dias_semana = (row.get("dias_semana") or "").strip()
                    class_obj.turma_label = (row.get("aluno_turma") or "").strip() or codigo

                    student = get_or_create_import_student(session, class_obj.id, aluno_nome)
                    if student:
                        counters["students_updated"] += 1
                    else:
                        student = models.ImportStudent(class_id=class_obj.id, nome=aluno_nome)
                        session.add(student)
                        counters["students_created"] += 1
            else:
                # Personal-only CSV — no class assignment
                student = get_or_create_import_student(session, None, aluno_nome)
                if student:
                    counters["students_updated"] += 1
                else:
                    student = models.ImportStudent(class_id=None, nome=aluno_nome)
                    session.add(student)
                    counters["students_created"] += 1

            student.whatsapp = _format_whatsapp(row.get("whatsapp"))
            student.data_nascimento = (row.get("data_nascimento") or "").strip()
            student.data_atestado = (row.get("data_atest") or "").strip()
            student.categoria = (row.get("categoria") or "").strip()
            student.genero = (row.get("genero") or "").strip()
            student.parq = (row.get("parq") or "").strip()
            student.atestado = parse_bool(row.get("atestado") or "")

        if apply_overrides:
            _apply_transfer_overrides(session)
        _dedupe_import_students(session)
        _dedupe_import_students_global(session)
        session.commit()
        _save_import_status(
            {
                "filename": file.filename,
                "last_import_at": datetime.utcnow().isoformat(),
                "rows_processed": rows_processed,
                **counters,
            }
        )
        return ImportResult(**counters)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"import error: {exc}")


@app.get("/api/import-data/status", response_model=ImportStatusOut)
def get_import_data_status() -> ImportStatusOut:
    payload = _load_import_status()
    return ImportStatusOut(**payload)

@app.get("/api/bootstrap", response_model=BootstrapOut)
def bootstrap(
    unit_id: Optional[int] = None,
    professor: Optional[str] = None,
    session: Session = Depends(get_session),
) -> BootstrapOut:
    removed = _dedupe_import_students(session) + _dedupe_import_students_global(session)
    if removed > 0:
        session.commit()

    uid_registry = _load_student_uid_registry()
    uid_registry_changed = False

    units_stmt = select(models.ImportUnit).order_by(models.ImportUnit.name)
    units = session.exec(units_stmt).all()

    classes_stmt = select(models.ImportClass)
    if unit_id is not None:
        classes_stmt = classes_stmt.where(models.ImportClass.unit_id == unit_id)
    classes_stmt = classes_stmt.order_by(models.ImportClass.codigo, models.ImportClass.horario)
    classes = session.exec(classes_stmt).all()

    # Mode switch by environment:
    # - unit (default): full unit visibility (legacy behavior / Bela Vista)
    # - professor: each professor sees only own classes/students
    if ACCESS_MODE == "professor":
        scoped_professor = _normalize_text_fold(professor)
        if scoped_professor:
            classes = [
                cls
                for cls in classes
                if _normalize_text_fold(cls.professor or "") == scoped_professor
            ]
        else:
            classes = []

    class_ids = [c.id for c in classes]
    students_stmt = select(models.ImportStudent)
    if class_ids:
        students_stmt = students_stmt.where(models.ImportStudent.class_id.in_(class_ids))
    students = list(session.exec(students_stmt).all())

    # Also include unallocated students (class_id IS NULL)
    unallocated_stmt = select(models.ImportStudent).where(models.ImportStudent.class_id.is_(None))
    unallocated = session.exec(unallocated_stmt).all()
    # Avoid duplicates if class_ids was empty (all students already fetched)
    if class_ids:
        students = students + list(unallocated)

    for student in students:
        _, changed = _ensure_student_uid_for_student(student, registry=uid_registry)
        uid_registry_changed = uid_registry_changed or changed

    if uid_registry_changed:
        _save_student_uid_registry(uid_registry)

    return BootstrapOut(
        units=[ImportUnitOut(id=u.id, name=u.name) for u in units],
        classes=[
            ImportClassOut(
                id=c.id,
                unit_id=c.unit_id,
                grupo=c.codigo,
                codigo=c.codigo,
                turma_label=c.turma_label,
                horario=c.horario,
                professor=c.professor,
                nivel=c.nivel,
                faixa_etaria=c.faixa_etaria,
                capacidade=c.capacidade,
                dias_semana=c.dias_semana,
            )
            for c in classes
        ],
        students=[
            _import_student_out(s, uid_registry=uid_registry)
            for s in students
        ],
    )

@app.post("/api/import-students", response_model=ImportStudentOut)
def create_import_student(payload: ImportStudentUpsertPayload, session: Session = Depends(get_session)) -> ImportStudentOut:
    nome = str(payload.nome or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="nome is required")

    has_class_info = any([str(payload.turma or "").strip(), str(payload.horario or "").strip(), str(payload.professor or "").strip()])

    if has_class_info:
        target_class = _find_import_class_by_triple(
            session=session,
            turma=payload.turma,
            horario=payload.horario,
            professor=payload.professor,
        )
        if not target_class:
            raise HTTPException(status_code=404, detail="Class not found for turma/horario/professor")
        target_class_id: Optional[int] = target_class.id
    else:
        target_class = None
        target_class_id = None

    student = get_or_create_import_student(session, target_class_id, nome)
    if not student:
        student = models.ImportStudent(class_id=target_class_id, nome=nome)
        session.add(student)

    student.class_id = target_class_id
    student.nome = nome
    student.whatsapp = _format_whatsapp(payload.whatsapp)
    student.data_nascimento = str(payload.data_nascimento or "").strip()
    student.data_atestado = str(payload.data_atestado or "").strip()
    student.categoria = str(payload.categoria or "").strip()
    student.genero = str(payload.genero or "").strip()
    student.parq = str(payload.parq or "").strip()
    student.atestado = bool(payload.atestado)

    session.commit()
    session.refresh(student)
    return _import_student_out(student, preferred_uid=payload.student_uid)

@app.put("/api/import-students/{student_id}", response_model=ImportStudentOut)
def update_import_student(student_id: int, payload: ImportStudentUpsertPayload, session: Session = Depends(get_session)) -> ImportStudentOut:
    student = session.get(models.ImportStudent, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Import student not found")

    nome = str(payload.nome or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="nome is required")

    previous_class_id = student.class_id
    has_class_info = any([str(payload.turma or "").strip(), str(payload.horario or "").strip(), str(payload.professor or "").strip()])

    if has_class_info:
        target_class = _find_import_class_by_triple(
            session=session,
            turma=payload.turma,
            horario=payload.horario,
            professor=payload.professor,
        )
        if not target_class:
            raise HTTPException(status_code=404, detail="Class not found for turma/horario/professor")
        target_class_id: Optional[int] = target_class.id
    else:
        target_class = None
        target_class_id = None

    existing_target = get_or_create_import_student(session, target_class_id, nome)
    target_student = student
    if existing_target and existing_target.id != student.id:
        target_student = existing_target
        session.delete(student)

    target_student.class_id = target_class_id
    target_student.nome = nome
    target_student.whatsapp = _format_whatsapp(payload.whatsapp)
    target_student.data_nascimento = str(payload.data_nascimento or "").strip()
    target_student.data_atestado = str(payload.data_atestado or "").strip()
    target_student.categoria = str(payload.categoria or "").strip()
    target_student.genero = str(payload.genero or "").strip()
    target_student.parq = str(payload.parq or "").strip()
    target_student.atestado = bool(payload.atestado)

    movement_type = str(payload.movement_type or "correction").strip().lower()
    is_transfer = movement_type == "transfer"
    if target_class is not None and previous_class_id != target_class_id:
        if is_transfer:
            _upsert_transfer_override_for_student(target_student, target_class)
        else:
            _remove_transfer_override_for_student(target_student)

    session.add(target_student)
    _dedupe_import_students(session)
    _dedupe_import_students_global(session)
    session.commit()
    session.refresh(target_student)
    return _import_student_out(target_student, preferred_uid=payload.student_uid)


@app.post("/api/import-students/bulk-allocate")
def bulk_allocate_import_students(
    payload: ImportStudentBulkAllocatePayload,
    session: Session = Depends(get_session),
):
    student_ids = sorted({int(student_id) for student_id in payload.student_ids if int(student_id) > 0})
    if not student_ids:
        raise HTTPException(status_code=400, detail="student_ids is required")

    target_class = _find_import_class_by_triple(
        session=session,
        turma=payload.turma,
        horario=payload.horario,
        professor=payload.professor,
    )
    if not target_class:
        raise HTTPException(status_code=404, detail="Class not found for turma/horario/professor")

    students = session.exec(
        select(models.ImportStudent).where(models.ImportStudent.id.in_(student_ids))
    ).all()
    found_ids = {int(student.id) for student in students if student.id is not None}
    missing_ids = [student_id for student_id in student_ids if student_id not in found_ids]
    if missing_ids:
        raise HTTPException(status_code=404, detail=f"Import students not found: {missing_ids}")

    movement_type = str(payload.movement_type or "correction").strip().lower()
    is_transfer = movement_type == "transfer"

    updated = 0
    for student in students:
        previous_class_id = student.class_id
        existing_target = get_or_create_import_student(session, target_class.id, student.nome)
        target_student = student
        if existing_target and existing_target.id != student.id:
            target_student = existing_target
            target_student.whatsapp = student.whatsapp
            target_student.data_nascimento = student.data_nascimento
            target_student.data_atestado = student.data_atestado
            target_student.categoria = student.categoria
            target_student.genero = student.genero
            target_student.parq = student.parq
            target_student.atestado = student.atestado
            session.delete(student)

        target_student.class_id = target_class.id
        if is_transfer and previous_class_id != target_class.id:
            _upsert_transfer_override_for_student(target_student, target_class)
        else:
            _remove_transfer_override_for_student(target_student)
        session.add(target_student)
        updated += 1

    _dedupe_import_students(session)
    _dedupe_import_students_global(session)
    session.commit()
    return {"ok": True, "updated": updated}

# Users endpoints (bootstrap)
@app.post("/users/register")
def register_user(username: str, password: str, session: Session = Depends(get_session)):
    stmt = select(models.User).where(models.User.username == username)
    existing = session.exec(stmt).first()
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")
    user = models.User(username=username, password_hash=get_password_hash(password), role="admin")
    session.add(user)
    session.commit()
    session.refresh(user)
    return {"username": user.username, "id": user.id}

@app.post("/token")
def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    unit_name: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    _validate_unit_for_environment(unit_name)
    user = authenticate_user(session, form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=400, detail="Incorrect username or password")
    access_token = create_access_token(data={"sub": user.username})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "unit_name": UNIT_NAME,
        "env_name": ENV_NAME,
    }

# Data endpoints
@app.get("/students", response_model=List[models.Student])
def list_students(limit: int = 100, session: Session = Depends(get_session)):
    return crud.get_students(session, limit)

@app.post("/students", response_model=models.Student)
def add_student(student: models.Student, session: Session = Depends(get_session)):
    return crud.create_student(session, student)

@app.delete("/students/{student_id}")
def delete_student(student_id: int, session: Session = Depends(get_session)):
    student = session.get(models.Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    session.delete(student)
    session.commit()
    return {"ok": True}

@app.get("/classes", response_model=List[models.ClassModel])
def list_classes(session: Session = Depends(get_session)):
    return crud.get_classes(session)

@app.post("/classes", response_model=models.ClassModel)
def add_class(class_model: models.ClassModel, session: Session = Depends(get_session)):
    return crud.create_class(session, class_model)

# Import classes endpoints (for turma management via UI)
@app.post("/import-classes")
def create_import_class(
    data: models.ImportClassCreate,
    session: Session = Depends(get_session),
):
    """Create a new import class (turma) - default unit_id is 1"""
    unit_id = 1  # Default unit
    # Generate codigo from turma components
    codigo = f"{data.professor[:2].lower()}{data.turma_label[:3].lower()}".replace(" ", "")
    
    new_class = models.ImportClass(
        unit_id=unit_id,
        codigo=codigo or f"t{session.exec(select(func.count(models.ImportClass.id))).first() or 0}",
        turma_label=data.turma_label,
        horario=data.horario,
        professor=data.professor,
        nivel=data.nivel or "",
        faixa_etaria=data.faixa_etaria or "",
        capacidade=data.capacidade or 0,
        dias_semana=data.dias_semana or "",
    )
    session.add(new_class)
    session.commit()
    session.refresh(new_class)
    return {
        "id": new_class.id,
        "unit_id": new_class.unit_id,
        "codigo": new_class.codigo,
        "turma_label": new_class.turma_label,
        "horario": new_class.horario,
        "professor": new_class.professor,
        "nivel": new_class.nivel,
        "faixa_etaria": new_class.faixa_etaria,
        "capacidade": new_class.capacidade,
        "dias_semana": new_class.dias_semana,
    }

@app.put("/import-classes/{class_id}")
def update_import_class(
    class_id: int,
    data: models.ImportClassUpdate,
    session: Session = Depends(get_session),
):
    """Update an existing import class (turma)"""
    import_class = session.exec(select(models.ImportClass).where(models.ImportClass.id == class_id)).first()
    if not import_class:
        raise HTTPException(status_code=404, detail="Class not found")
    
    if data.turma_label is not None:
        import_class.turma_label = data.turma_label
    if data.horario is not None:
        import_class.horario = data.horario
    if data.professor is not None:
        import_class.professor = data.professor
    if data.nivel is not None:
        import_class.nivel = data.nivel
    if data.faixa_etaria is not None:
        import_class.faixa_etaria = data.faixa_etaria
    if data.capacidade is not None:
        import_class.capacidade = data.capacidade
    if data.dias_semana is not None:
        import_class.dias_semana = data.dias_semana
    
    session.add(import_class)
    session.commit()
    session.refresh(import_class)
    return {
        "id": import_class.id,
        "unit_id": import_class.unit_id,
        "codigo": import_class.codigo,
        "turma_label": import_class.turma_label,
        "horario": import_class.horario,
        "professor": import_class.professor,
        "nivel": import_class.nivel,
        "faixa_etaria": import_class.faixa_etaria,
        "capacidade": import_class.capacidade,
        "dias_semana": import_class.dias_semana,
    }

@app.get("/attendance", response_model=List[models.Attendance])
def list_attendance(session: Session = Depends(get_session)):
    statement = session.exec(select(models.Attendance))
    return statement.all()

@app.post("/attendance", response_model=models.Attendance)
def add_attendance(attendance: models.Attendance, session: Session = Depends(get_session)):
    return crud.create_attendance(session, attendance)

# Import endpoint (protected) - accepts uploaded file or file param (file must be in data/)
@app.post("/import")
def import_excel(
    file: Optional[str] = Query(None, description="Nome do arquivo em /app/data/"),
    out_clean: Optional[bool] = Query(False, description="Salvar arquivo cleaned em data/"),
    upload: Optional[UploadFile] = File(None),
    session: Session = Depends(get_session),
    current_user: models.User = Depends(get_current_user),
):
    os.makedirs(DATA_DIR, exist_ok=True)

    if upload:
        dest_path = os.path.join(DATA_DIR, upload.filename)
        with open(dest_path, "wb") as f:
            f.write(upload.file.read())
        file_path = dest_path
    elif file:
        file_path = os.path.join(DATA_DIR, file)
    else:
        raise HTTPException(status_code=400, detail="No file provided")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    out_cleaned_path = None
    if out_clean:
        base = os.path.basename(file_path)
        out_cleaned_path = os.path.join(DATA_DIR, f"{base}.cleaned.xlsx")

    result = import_from_excel(file_path, session, out_cleaned=out_cleaned_path)
    response = {"imported": result["counts"], "mapping": result["mapping"]}
    if result.get("cleaned_path"):
        response["cleaned_path"] = os.path.basename(result["cleaned_path"])
    return response

# File download endpoint
@app.get("/files/{filename}")
def download_file(filename: str):
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
